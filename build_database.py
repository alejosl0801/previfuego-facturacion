#!/usr/bin/env python3
"""
build_database.py  –  Builds BASE_DATOS_KFC.xlsx (198 locales)

Sources (Dropbox):
  1. PRESUPUESTO PROVEEDORES 2026  → brand names (source of truth)
  2. MATRIZ EXTINTORES GRUPO KFC   → extintor data (primary)
  3. PROYECCIÓN INGRESOS MENSUAL 2026 → dates + data for missing locals

Logic:
  - 200 PRESUPUESTO PREVIFUEGO rows
  - Remove G042 (closed) → 199
  - H068 merges into K172 (not a separate local) → 198 unique locals
  - V091 not yet in PRESUPUESTO but IS in PROYECCIÓN → counted as the 198th
  - ALL locals are in OCT2026-SEP2027 recarga cycle
  - COBRO = RECARGA price (which already includes mantt)
  - Cycle sheet months: OCT/NOV/DEC → recarga 2026; JAN-SEP → recarga 2027
"""

import requests, io, re, time, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict, Counter

from dropbox_auth import get_access_token
TOKEN = get_access_token()

FILES = {
    "PRESUPUESTO": "/Previfuego/2026/PRESUPUESTO PROVEEDORES AÑ0 2026.xlsx",
    "MATRIZ":      "/Previfuego/MATRIZ LOCALES/MATRIZ EXTINTORES GRUPO KFC.xlsx",
    "PROYECCION":  "/Previfuego/PRESUPUESTOS/PROYECCION INGRESOS MENSUAL 2026.xlsx",
}

MONTHS = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
          "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
MONTH_IDX = {m: i for i, m in enumerate(MONTHS)}
OCT_TO_DEC = {"OCTUBRE", "NOVIEMBRE", "DICIEMBRE"}

# Locales fuera de Guayaquil: movilización/viáticos extraídos de cotizaciones 2026
# ckey → (descripción ruta, valor $)
VIATICOS = {
    ('K', 79):  ('GYE-BABAHOYO 75KM',         39.00),
    ('K', 65):  ('GYE-MILAGRO 50KM',           25.00),
    ('V', 75):  ('GYE-MACHALA',                 8.00),
    ('V', 79):  ('GYE-MILAGRO',                 8.00),
    ('K', 90):  ('GYE-MACHALA 185KM',          55.50),
    ('K', 91):  ('GYE-MACHALA 185KM',          55.50),
    ('K', 148): ('GYE-MACHALA 185KM',          55.50),
    ('K', 82):  ('GYE-PLAYAS',                 44.00),
    ('J', 20):  ('GYE-LIBERTAD 136KM x2',        68.00),  # RETIRO+ENTREGA
    ('K', 93):  ('GYE-QUEVEDO',                35.00),
    ('K', 106): ('GYE-QUEVEDO',                35.00),
    ('M', 32):  ('GYE-QUEVEDO',                35.00),
    ('K', 149): ('GYE-QUEVEDO/VENTANAS',       35.00),
    ('K', 78):  ('GYE-DAULE',                  25.00),
    ('M', 39):  ('GYE-BABAHOYO',               24.00),
    ('K', 53):  ('GYE-BABAHOYO',               24.00),
    ('R', 10):  ('GYE-ORELLANA',               18.80),
}

# Locales con AÑO_RECARGA=2026 aunque su mes de servicio sea ENE-SEP
# (su recarga ya ocurrió en el ciclo OCT2025-SEP2026)
RECARGA_2026_OVERRIDE = {('R', 3), ('R', 8), ('R', 10)}

# ── Correcciones manuales (revisión cotizaciones 2026) ─────────────────────────
# El usuario confirmó el inventario real de extintores por local.
# ckey → lista de (tipo, capacidad, cantidad). Reemplaza lo que venga de MATRIZ.
EXTINTOR_OVERRIDE = {
    # ── ENERO ──
    ('K', 25):  [('CO2','75',1), ('CO2','10',1), ('PQS','10',3)],                                  # Mobil Durán      = $46
    ('K', 154): [('CO2','50',1), ('CO2','5',3), ('PQS','20',1), ('PQS','10',3), ('TIPO K','2.5',1)], # Pascuales       = $54
    ('K', 76):  [('CO2','50',1), ('CO2','5',2), ('PQS','10',4)],                                    # Parque Centenario= $40
    ('K', 105): [('CO2','75',1), ('CO2','5',1), ('PQS','20',1), ('PQS','10',1)],                    # Hiper Market    = $44
    # ── FEBRERO ──
    ('K', 45):  [('CO2','50',1), ('CO2','5',1), ('PQS','20',3), ('PQS','10',6)],                    # Boyacá          = $70
    # ── MARZO ──
    ('K', 75):  [('CO2','50',1), ('CO2','10',1), ('PQS','20',1), ('PQS','10',2)],                   # Mobil Kennedy   = $40
    ('K', 94):  [('CO2','50',1), ('CO2','10',1), ('CO2','5',2), ('PQS','10',1)],                    # Outlet Durán    = $32
    ('M', 11):  [('CO2','75',1), ('PQS','10',2), ('CO2','5',1), ('TIPO K','2.5',1)],                # Mall del Sur    = $48
    ('K', 145): [('CO2','50',1), ('CO2','5',3), ('PQS','20',1), ('PQS','10',2), ('TIPO K','2.5',1)],# Aki San Eduardo = $50
    ('M', 37):  [('CO2','50',1), ('CO2','5',2), ('PQS','10',4), ('TIPO K','2.5',1)],                # 9 de Octubre    = $48
    # ── MAYO ──  (R004: recarga fue mayo 2025, mantenimiento mayo 2026 → valor MANTT)
    ('R', 4):   [('CO2','50',2), ('CO2','5',3), ('PQS','10',2), ('TIPO K','2.5',1)],                # San Marino mantt= $62
    ('R', 10):  [('CO2','75',1), ('CO2','5',4), ('CO2','10',1), ('PQS','20',2), ('PQS','10',10), ('TIPO K','2.5',2)],  # Orellana (recarga+mantt mayo 2026)
    ('K', 96):  [('CO2','50',2), ('CO2','5',4), ('PQS','20',2), ('PQS','10',3), ('TIPO K','2.5',1)],# Portete         = $84
    # ── JUNIO ──
    ('F', 3):   [('CO2','50',1), ('CO2','5',1), ('PQS','10',2), ('TIPO K','2.5',1)],                # Dolce Incontro AER = $38
    ('J', 8):   [('CO2','75',1), ('PQS','10',2)],                                                   # Terminal        = $38
    ('K', 83):  [('CO2','75',1), ('CO2','5',2), ('PQS','20',2), ('PQS','10',1)],                    # Mucho Lote      = $54
    ('K', 79):  [('CO2','50',1), ('CO2','10',2), ('PQS','10',2)],                                   # Babahoyo (mes real JUNIO) ext = $36 + viáticos
    # ── JULIO ──
    ('K', 162): [('CO2','50',1), ('CO2','5',3), ('PQS','20',1), ('PQS','10',2), ('TIPO K','2.5',1)],# Daule Estación  = $50
    ('K', 171): [('CO2','50',1), ('CO2','10',1), ('PQS','10',2), ('TIPO K','2.5',1)],               # Shell Abu Dabi  = $40
    # ── OCTUBRE (recarga) ──
    ('K', 88):  [('CO2','50',1), ('CO2','5',3), ('PQS','20',1), ('PQS','10',2)],                    # Hipermarket Vía Daule recarga = $97.70
    ('K', 110): [('CO2','75',1), ('CO2','10',1), ('CO2','5',2), ('PQS','20',1), ('PQS','10',2)],    # City Mall PB recarga = $124.70
    ('B', 1):   [('CO2','50',1), ('CO2','20',1), ('PQS','10',3), ('PQS','5',2)],                    # Sports Bar
}

# Extintores con servicio mixto en el mismo mes: (tipo, cap) listados aquí reciben
# solo MANTENIMIENTO aunque el mes sea de RECARGA (su recarga ya se hizo antes).
# El resto de extintores del mismo local sigue el tipo de servicio normal.
MIXED_SERVICE_MANTT = {
    ('B', 1): {('CO2', '50')},   # B1 Sports Bar OCT: 50CO2 ya recargó; solo mantt
}

# Locales que ya no se atienden → eliminar por completo
DELETE_LOCALS = {('K', 124)}   # Gran Aki Loja – ya no corresponde

# Mes de servicio corregido (cotización en otra fecha era una venta, no mantenimiento)
MES_OVERRIDE = {
    ('K', 79): 'JUNIO',     # cotización ENERO = venta; mantenimiento real JUNIO
    ('K', 88): 'OCTUBRE',   # cotización ABRIL = venta; recarga real OCTUBRE
}

# Años de recarga especiales (ckey → (AÑO_ULT_RECARGA, AÑO_RECARGA))
ANO_OVERRIDE = {
    ('R', 4): (2025, 2028),   # recargó mayo 2025, próxima recarga mayo 2028 (2026 = mantenimiento)
}

# ── Pricing ───────────────────────────────────────────────────────────────────

MANTT = {
    ("PQS",    "5"):   2.00,
    ("PQS",    "10"):  4.00,
    ("PQS",    "20"):  8.00,
    ("CO2",    "5"):   2.00,
    ("CO2",    "10"):  4.00,
    ("CO2",    "20"):  8.00,
    ("CO2",    "50"):  20.00,
    ("CO2",    "75"):  30.00,
    ("TIPO K", "2.5"): 8.00,
}
RECARGA = {
    ("PQS",    "5"):   4.90,
    ("PQS",    "10"):  9.80,
    ("PQS",    "20"):  19.60,
    ("CO2",    "5"):   4.50,
    ("CO2",    "10"):  9.00,
    ("CO2",    "20"):  18.00,
    ("CO2",    "50"):  45.00,
    ("CO2",    "75"):  67.50,
    ("TIPO K", "2.5"): 98.80,
}
CAP_DISPLAY = {
    ("PQS",    "5"):   "5 LBS",
    ("PQS",    "10"):  "10 LBS",
    ("PQS",    "20"):  "20 LBS",
    ("CO2",    "5"):   "5 LBS",
    ("CO2",    "10"):  "10 LBS",
    ("CO2",    "20"):  "20 LBS",
    ("CO2",    "50"):  "50 LBS",
    ("CO2",    "75"):  "75 LBS",
    ("TIPO K", "2.5"): "2.5 GLS",
}

# Brand fallback when not in PRESUPUESTO (for new locals K192, K194, V091, etc.)
BRAND_FALLBACK = {
    "K":  "KFC",
    "G":  "GUS",
    "M":  "MENESTRAS DEL NEGRO",
    "T":  "TROPIBURGER",
    "V":  "JUAN VALDEZ",
    "I":  "ILCAPPO",
    "J":  "CAJUN",
    "BS": "BASKIN ROBBINS",
    "CN": "CINNABON",
    "DI": "DOLCE INCONTRO",
    "CA": "CAFA",
    "A":  "AMERICAN DELI",
    "B":  "AMERICAN DELI",
    "F":  "AMERICAN DELI",
    "E":  "ESPAÑOL",
    "R":  "CASA RES",
    "H":  "HELADERIAS",
}

# Normalise brand names from PRESUPUESTO to consistent forms
MARCA_NORM = {
    "BASKIN":             "BASKIN ROBBINS",
    "DOLCE":              "DOLCE INCONTRO",
    "CAFA":               "CAFA",
    "GUS":                "GUS",
    "ILCAPPO":            "IL CAPPO",
    "HELADERIAS":         "HELADERIAS",
    "CASA RES":           "CASA RES",
    "MENESTRAS DEL NEGRO":"MENESTRAS DEL NEGRO",
}


def marca_norm(raw):
    if not raw: return ""
    s = str(raw).strip()
    return MARCA_NORM.get(s, s)


# ── Dropbox download ──────────────────────────────────────────────────────────

def dropbox_download(path, retries=4):
    delay = 2
    for attempt in range(retries + 1):
        try:
            r = requests.post(
                "https://content.dropboxapi.com/2/files/download",
                headers={
                    "Authorization": f"Bearer {TOKEN}",
                    "Dropbox-API-Arg": json.dumps({"path": path}),
                },
                timeout=90,
            )
            if r.status_code in (503, 429) and attempt < retries:
                print(f"  HTTP {r.status_code} – retry in {delay}s …")
                time.sleep(delay); delay *= 2; continue
            r.raise_for_status()
            return r.content
        except requests.RequestException as e:
            if attempt < retries:
                print(f"  Error: {e} – retry in {delay}s …")
                time.sleep(delay); delay *= 2
            else:
                raise


# ── Canonical CC key ──────────────────────────────────────────────────────────

def canonical(raw):
    """Return (prefix, number) tuple, or None if not parseable.
    Prefix is the canonical brand prefix (K, G, V, J, CN, BS, …).
    """
    if not raw:
        return None
    s = str(raw).strip().upper()
    s = re.sub(r'EC$', '', s).strip()

    for pat, pfx in [
        (r'^KFC(?:K+)?\s*(\d+)',  'K'),   # KFC02, KFCK192
        (r'^GUS\s*(\d+)',          'G'),   # GUS49
        (r'^JV\s*(\d+)',           'V'),   # JV15 → Juan Valdez
        (r'^JB\s*(\d+)',           'V'),   # JB = alternate Juan Valdez notation
        (r'^CJNC\s*(\d+)',         'J'),   # CJNC31 = CAJUN in MATRIZ
        (r'^CN\s*(\d+)',           'CN'),  # CN04 = CINNABON
        (r'^BS\s*(\d+)',           'BS'),  # BS04 = BASKIN ROBBINS
        (r'^DI\s*(\d+)',           'DI'),  # DI01 = DOLCE INCONTRO
        (r'^CA\s*(\d+)',           'CA'),  # CA01 = CAFA
        (r'^F\s*0*(\d+)',          'F'),   # F003 = AMERICAN DELI (Aeropuerto)
    ]:
        m = re.match(pat, s)
        if m:
            return (pfx, int(m.group(1)))

    # Single letter + optional space + digits
    m = re.match(r'^([A-Z])\s*(\d+)', s)
    if m:
        pfx, num = m.group(1), int(m.group(2))
        # J alone (not JV/JB) = CAJUN per user
        return (pfx, num)

    # Multi-letter + digits (handles ILCI02, AMCA19, ESPE07, TROT54 etc.)
    m = re.match(r'^([A-Z]{2,4})\s*(\d+)', s)
    if m:
        pfx, num = m.group(1), int(m.group(2))
        return (pfx, num)

    return None


def fmt_code(pfx, num):
    """Display code from canonical pair."""
    return f"{pfx}{num:03d}"


# ── Tipo / capacidad normalisation ────────────────────────────────────────────

def norm_tipo(raw):
    if raw is None: return ""
    s = str(raw).strip().upper()
    if re.match(r'^(K|TIPO\s*K|TYPE\s*K)$', s): return "TIPO K"
    if s.startswith("CO2"): return "CO2"
    if s.startswith("PQS"): return "PQS"
    return s


def norm_cap(raw):
    if raw is None: return ""
    if isinstance(raw, (int, float)):
        v = float(raw)
        return str(int(v)) if v == int(v) else str(round(v, 1))
    s = str(raw).strip().upper()
    s = re.sub(r'[A-Z\s]', '', s)
    s = s.replace(",,", ".").replace(",", ".")
    s = re.sub(r'\.{2,}', '.', s)
    m = re.match(r'(\d+(?:\.\d+)?)', s)
    if m:
        v = float(m.group(1))
        return str(int(v)) if v == int(v) else str(round(v, 1))
    return ""


def norm_tipo_cap(tipo_r, cap_r):
    tipo = norm_tipo(tipo_r)
    cap  = norm_cap(cap_r)
    # PQS 2.5 doesn't exist → TIPO K 2.5 GLS
    if tipo == "PQS" and cap == "2.5":
        tipo = "TIPO K"
    # Cap raw has 'G' (gallons) and value is 2.5 → TIPO K
    if cap == "2.5" and cap_r and re.search(r'G', str(cap_r).upper()):
        tipo = "TIPO K"
    return tipo, cap


def to_year(v):
    if v is None: return ""
    try: return str(int(float(str(v))))
    except: return str(v).strip()


# ── Download ──────────────────────────────────────────────────────────────────

print("Downloading PRESUPUESTO …")
data_pres = dropbox_download(FILES["PRESUPUESTO"])
print(f"  {len(data_pres):,} bytes")

print("Downloading MATRIZ …")
data_mat = dropbox_download(FILES["MATRIZ"])
print(f"  {len(data_mat):,} bytes")

print("Downloading PROYECCION …")
data_proy = dropbox_download(FILES["PROYECCION"])
print(f"  {len(data_proy):,} bytes")

wb_pres = openpyxl.load_workbook(io.BytesIO(data_pres), data_only=True)
wb_mat  = openpyxl.load_workbook(io.BytesIO(data_mat),  data_only=True)
wb_proy = openpyxl.load_workbook(io.BytesIO(data_proy), data_only=True)


# ── 1. Parse PRESUPUESTO (sheet SSO) ──────────────────────────────────────────
# Structure: header at row 4 (0-indexed)
#   col 1=EMPRESA, col 2=LOCAL (CC), col 3=C.COSTO, col 4=MARCA,
#   col 5=NOMBRE LOCAL, col 6=RESPONSABLE, col 7=Proveedor

print("\n=== Parsing PRESUPUESTO ===")
presupuesto_map = {}  # canonical_key -> {marca, nombre, cc_pres}

ws_sso = wb_pres["SSO"]
pres_rows = list(ws_sso.iter_rows(values_only=True))

for row in pres_rows[5:]:  # skip title + blank + header rows
    if not row or len(row) < 8:
        continue
    proveedor = str(row[7]).strip() if row[7] else ""
    if proveedor.upper() != "PREVIFUEGO":
        continue
    cc_raw = row[2]
    marca  = str(row[4]).strip() if row[4] else ""
    nombre = str(row[5]).strip() if row[5] else ""
    ckey = canonical(cc_raw)
    if not ckey:
        continue
    presupuesto_map[ckey] = {
        "marca":   marca_norm(marca),
        "nombre":  nombre,
        "cc_pres": str(cc_raw).strip() if cc_raw else "",
    }

print(f"PRESUPUESTO entries: {len(presupuesto_map)}")

# Confirm G042 and H068 are there
G042_KEY = ('G', 42)
H068_KEY = ('H', 68)
K172_KEY = ('K', 172)
print(f"  G042 in PRESUPUESTO: {G042_KEY in presupuesto_map}")
print(f"  H068 in PRESUPUESTO: {H068_KEY in presupuesto_map}")


# ── 2. Parse MATRIZ and PROYECCIÓN ────────────────────────────────────────────

SKIP_TIPO = re.compile(r'^(MANTENIMIENTO|SIST|SISTEMA|TOTAL|SUBTOTAL)', re.I)

def parse_extintor_sheet(ws, mes):
    """Return list of extintor dicts from a monthly sheet."""
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row (first row with ≥2 header-like keywords)
    header_idx = 0
    for i, row in enumerate(all_rows[:10]):
        cells = [str(c).strip().upper() if c else "" for c in row]
        hits = sum(1 for c in cells if c in ("CC","TIPO","CAPACIDAD","UBICACION","UBICACIÓN","LOCAL","NOMBRE"))
        if hits >= 2:
            header_idx = i
            break

    hrow = [str(c).strip().upper() if c else "" for c in all_rows[header_idx]]

    def find_col(*names):
        for name in names:
            for j, h in enumerate(hrow):
                if name in h:
                    return j
        return None

    c_cc   = find_col("CC") or 0
    c_ubic = find_col("UBICACION", "UBICACIÓN", "NOMBRE", "LOCAL") or 1
    c_tipo = find_col("TIPO") or 2
    c_cap  = find_col("CAPACIDAD", "CAP") or 3
    c_ult  = find_col("ULT", "ÚLTIMA", "ULTIMA", "ANTERIOR") or 9
    c_prox = find_col("PRÓX", "PROX", "SIGUIENTE") or 10

    rows_out = []
    cur_ckey   = None
    cur_cc_raw = None

    for row in all_rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        def gc(idx):
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        cc_raw  = row[c_cc] if c_cc < len(row) else None
        ubic    = gc(c_ubic)
        tipo_r  = row[c_tipo] if c_tipo < len(row) else None
        cap_r   = row[c_cap]  if c_cap  < len(row) else None
        ult_r   = row[c_ult]  if c_ult  < len(row) else None
        prox_r  = row[c_prox] if c_prox < len(row) else None

        # Forward-fill CC
        if cc_raw is not None and str(cc_raw).strip():
            ckey = canonical(str(cc_raw).strip())
            if ckey:
                cur_ckey   = ckey
                cur_cc_raw = str(cc_raw).strip()

        if not cur_ckey:
            continue

        tipo, cap = norm_tipo_cap(tipo_r, cap_r)

        if not tipo or tipo not in ("PQS", "CO2", "TIPO K"):
            continue
        if tipo_r and SKIP_TIPO.match(str(tipo_r)):
            continue

        rows_out.append({
            "ckey":    cur_ckey,
            "cc_raw":  cur_cc_raw,
            "mes":     mes,
            "ubic":    ubic,
            "tipo":    tipo,
            "cap":     cap,
            "año_ult": to_year(ult_r),
            "año_prox": to_year(prox_r),
        })

    return rows_out


print("\n=== Parsing MATRIZ ===")
matriz_rows = []
for sname in wb_mat.sheetnames:
    mes = sname.strip().upper()
    if mes not in MONTHS:
        continue
    rows = parse_extintor_sheet(wb_mat[sname], mes)
    print(f"  {sname}: {len(rows)} extintor rows")
    matriz_rows.extend(rows)

print(f"\nTotal MATRIZ extintor rows: {len(matriz_rows)}")
mat_keys = set(r["ckey"] for r in matriz_rows)
print(f"Unique MATRIZ locals:       {len(mat_keys)}")


print("\n=== Parsing PROYECCIÓN ===")
proy_rows = []
for sname in wb_proy.sheetnames:
    mes = sname.strip().upper()
    if mes not in MONTHS:
        continue
    rows = parse_extintor_sheet(wb_proy[sname], mes)
    print(f"  {sname}: {len(rows)} extintor rows")
    proy_rows.extend(rows)

proy_keys = set(r["ckey"] for r in proy_rows)
print(f"\nTotal PROYECCIÓN extintor rows: {len(proy_rows)}")
print(f"Unique PROYECCIÓN locals:       {len(proy_keys)}")

only_proy = proy_keys - mat_keys
print(f"\nIn PROYECCIÓN but NOT MATRIZ: {sorted(only_proy)}")


# ── 3. Apply corrections ──────────────────────────────────────────────────────

print("\n=== Applying corrections ===")

# (a) Remove G042 (closed)
before = len(matriz_rows)
matriz_rows = [r for r in matriz_rows if r["ckey"] != G042_KEY]
print(f"Removed G042: {before - len(matriz_rows)} rows removed")

# (b) H068 → already absent from MATRIZ (its extintores were merged into K172
#     which already includes 'heladería' extintor; no additional rows needed)
h068_in_mat = [r for r in matriz_rows if r["ckey"] == H068_KEY]
print(f"H068 rows in MATRIZ: {len(h068_in_mat)} (expected 0 – already part of K172)")

# (c) Add any missing locals from PROYECCIÓN (generalised: all missing, not just V091)
# SHARED_BS_CN defined later in step 4; these 4 CNs are already covered by their BS equivalents
SHARED_CN = {('CN', 4), ('CN', 16), ('CN', 31), ('CN', 37)}
pres_expected = set(
    k for k in presupuesto_map
    if k not in (G042_KEY, H068_KEY) and k not in SHARED_CN and k not in DELETE_LOCALS
)
mat_keys_now = set(r["ckey"] for r in matriz_rows)
proy_by_key = {}
for r in proy_rows:
    proy_by_key.setdefault(r["ckey"], []).append(r)

added_from_proy = []
truly_missing = []

for ckey in sorted(pres_expected - mat_keys_now):
    if ckey in proy_by_key:
        seen = set()
        deduped = []
        for r in proy_by_key[ckey]:
            sig = (r["tipo"], r["cap"], r["ubic"])
            if sig not in seen:
                seen.add(sig); deduped.append(r)
        matriz_rows.extend(deduped)
        added_from_proy.append((ckey, len(deduped)))
    else:
        truly_missing.append(ckey)

# Also ensure V091 is included even if already in mat_keys (backwards compat)
V091_KEY = ('V', 91)
if V091_KEY not in mat_keys_now and V091_KEY not in [c for c, _ in added_from_proy]:
    if V091_KEY in proy_by_key:
        seen = set(); deduped = []
        for r in proy_by_key[V091_KEY]:
            sig = (r["tipo"], r["cap"], r["ubic"])
            if sig not in seen:
                seen.add(sig); deduped.append(r)
        matriz_rows.extend(deduped)
        added_from_proy.append((V091_KEY, len(deduped)))

if added_from_proy:
    print(f"\nAdded from PROYECCIÓN ({len(added_from_proy)} locals):")
    for ckey, n in added_from_proy:
        info = presupuesto_map.get(ckey, {})
        print(f"  {fmt_code(*ckey)}: {info.get('nombre','?')} – {n} extintores")
else:
    print("\nNo missing PRESUPUESTO locals (all already in MATRIZ)")

if truly_missing:
    print(f"\n*** TRULY MISSING (not in MATRIZ nor PROYECCIÓN) – {len(truly_missing)} locals: ***")
    for ckey in truly_missing:
        info = presupuesto_map[ckey]
        print(f"  {fmt_code(*ckey)}: {info['cc_pres']} – {info['marca']} – {info['nombre']}")
else:
    print("All PRESUPUESTO locals resolved ✓")

# (d) Override MES from PROYECCIÓN where PROYECCIÓN is source of truth
proy_mes = {}
for r in proy_rows:
    if r["ckey"] not in proy_mes:
        proy_mes[r["ckey"]] = r["mes"]

overrides = 0
for r in matriz_rows:
    pm = proy_mes.get(r["ckey"])
    if pm and pm != r["mes"]:
        r["mes"] = pm
        overrides += 1
print(f"\nMES overridden from PROYECCIÓN: {overrides} rows")

# (e) Correcciones manuales según revisión de cotizaciones 2026 ────────────────
print("\n=== Aplicando correcciones manuales (cotizaciones 2026) ===")

# Eliminar locales que ya no se atienden
before = len(matriz_rows)
matriz_rows = [r for r in matriz_rows if r["ckey"] not in DELETE_LOCALS]
print(f"Locales eliminados {sorted(DELETE_LOCALS)}: {before - len(matriz_rows)} filas removidas")

# Reemplazar listas de extintores corregidas
for ckey, spec in EXTINTOR_OVERRIDE.items():
    existing = [r for r in matriz_rows if r["ckey"] == ckey]
    ubic   = existing[0]["ubic"]   if existing else fmt_code(*ckey)
    cc_raw = existing[0]["cc_raw"] if existing else fmt_code(*ckey)
    mes    = MES_OVERRIDE.get(ckey) or (existing[0]["mes"] if existing else "")
    matriz_rows = [r for r in matriz_rows if r["ckey"] != ckey]
    for tipo, cap, qty in spec:
        for _ in range(qty):
            matriz_rows.append({
                "ckey": ckey, "cc_raw": cc_raw, "mes": mes, "ubic": ubic,
                "tipo": tipo, "cap": cap, "año_ult": "", "año_prox": "",
            })
    total_mantt = sum(MANTT.get((t, c), 0) * q for t, c, q in spec)
    n_ext = sum(q for _, _, q in spec)
    flag = "" if existing else "  [NUEVO - no estaba en MATRIZ]"
    print(f"  {fmt_code(*ckey)}: {n_ext} ext, mantt=${total_mantt:.2f}, mes={mes}{flag}")

# Aplicar mes corregido a locales NO incluidos en EXTINTOR_OVERRIDE
for ckey, mes in MES_OVERRIDE.items():
    if ckey in EXTINTOR_OVERRIDE:
        continue
    for r in matriz_rows:
        if r["ckey"] == ckey:
            r["mes"] = mes

# Final unique locals
final_keys = set(r["ckey"] for r in matriz_rows)
print(f"\nFinal unique locals: {len(final_keys)}")


# ── 4. Brand lookup ───────────────────────────────────────────────────────────

# Ubicaciones compartidas Baskin Robbins / Cinnabon (mismo local físico).
# En MATRIZ se codifican como BS; en PRESUPUESTO algunas aparecen como CN.
# Mapea la clave BS -> clave CN equivalente en PRESUPUESTO.
SHARED_BS_CN = {
    ('BS', 4):  ('CN', 4),
    ('BS', 16): ('CN', 16),
    ('BS', 37): ('CN', 37),
    ('BS', 31): ('CN', 31),
}

def get_marca(ckey):
    # Ubicación compartida Baskin/Cinnabon
    if ckey in SHARED_BS_CN:
        return "BASKIN ROBBINS / CINNABON"
    pinfo = presupuesto_map.get(ckey)
    if pinfo and pinfo["marca"]:
        return pinfo["marca"]
    pfx = ckey[0]
    return BRAND_FALLBACK.get(pfx, BRAND_FALLBACK.get(pfx[:1], "DESCONOCIDO"))


def get_nombre_local(ckey):
    pinfo = presupuesto_map.get(ckey)
    if pinfo and pinfo["nombre"]:
        return pinfo["nombre"]
    # Para ubicaciones compartidas, tomar el nombre de la entrada CN del PRESUPUESTO
    if ckey in SHARED_BS_CN:
        cn = presupuesto_map.get(SHARED_BS_CN[ckey])
        if cn and cn["nombre"]:
            return cn["nombre"]
    return ""


# ── 5. Build final rows ───────────────────────────────────────────────────────

print("\n=== Building final rows ===")

final_rows = []
for r in matriz_rows:
    ckey  = r["ckey"]
    mes   = r["mes"]
    tipo  = r["tipo"]
    cap   = r["cap"]

    pk  = (tipo, cap)
    cm  = MANTT.get(pk)
    cr  = RECARGA.get(pk)
    cd  = CAP_DISPLAY.get(pk, f"{cap}")

    # COBRO = RECARGA (includes mantt). If price unknown, fall back to MANTT.
    # Exception: mixed-service extintores get MANTT only.
    if ckey in MIXED_SERVICE_MANTT and (tipo, cap) in MIXED_SERVICE_MANTT[ckey]:
        cobro      = cm if cm is not None else None
        cobro_tipo = "MANTT (parcial — recarga previa)"
    else:
        cobro      = cr if cr is not None else (cm if cm is not None else None)
        cobro_tipo = "RECARGA (incl. mantt)" if cr is not None else ("MANTT" if cm is not None else "SIN PRECIO")

    # Años de recarga: override manual > R003/R008/R010 (mayo 2026) > regla general
    if ckey in ANO_OVERRIDE:
        ano_ult_rec, ano_rec = ANO_OVERRIDE[ckey]
    elif ckey in RECARGA_2026_OVERRIDE:
        ano_rec = 2026
        ano_ult_rec = ano_rec - 3
    else:
        ano_rec = 2026 if mes in OCT_TO_DEC else 2027
        ano_ult_rec = ano_rec - 3

    marca  = get_marca(ckey)
    nombre = get_nombre_local(ckey)
    codigo = fmt_code(*ckey)

    notas = []
    if cm is None:
        notas.append(f"precio no hallado ({tipo}/{cap})")

    final_rows.append({
        "CÓDIGO":          codigo,
        "CC_ORIGINAL":     r.get("cc_raw", ""),
        "MARCA":           marca,
        "NOMBRE_LOCAL":    nombre,
        "MES_SERVICIO":    mes,
        "UBICACIÓN":       r.get("ubic", ""),
        "TIPO":            tipo,
        "CAPACIDAD":       cd,
        "COSTO_MANTT":     round(cm, 2) if cm is not None else None,
        "PRECIO_RECARGA":  round(cr, 2) if cr is not None else None,
        "COBRO_ANUAL_EXT": round(cobro, 2) if cobro is not None else None,
        "TIPO_COBRO":      cobro_tipo,
        "AÑO_RECARGA":     ano_rec,
        "AÑO_ULT_RECARGA": ano_ult_rec,
        "NOTAS":           "; ".join(notas),
        "_cm":   cm   or 0.0,
        "_cr":   cr   or 0.0,
        "_cobro": cobro or 0.0,
        "_ckey": ckey,
    })

print(f"Total extintor rows: {len(final_rows)}")
print(f"Unique locals:       {len(set(r['CÓDIGO'] for r in final_rows))}")

# ── Inject viáticos/movilización rows ─────────────────────────────────────────
viat_rows = []
for r in final_rows:
    ckey = r["_ckey"]
    if ckey in VIATICOS and ckey not in {rv["_ckey"] for rv in viat_rows}:
        ruta, monto = VIATICOS[ckey]
        viat_rows.append({
            "CÓDIGO":          r["CÓDIGO"],
            "CC_ORIGINAL":     r["CC_ORIGINAL"],
            "MARCA":           r["MARCA"],
            "NOMBRE_LOCAL":    r["NOMBRE_LOCAL"],
            "MES_SERVICIO":    r["MES_SERVICIO"],
            "UBICACIÓN":       ruta,
            "TIPO":            "MOVILIZACIÓN",
            "CAPACIDAD":       "",
            "COSTO_MANTT":     monto,
            "PRECIO_RECARGA":  monto,
            "COBRO_ANUAL_EXT": monto,
            "TIPO_COBRO":      "MOVILIZACIÓN",
            "AÑO_RECARGA":     None,
            "AÑO_ULT_RECARGA": None,
            "NOTAS":           "Viáticos/movilización según cotización",
            "_cm":   monto,
            "_cr":   monto,
            "_cobro": monto,
            "_ckey": ckey,
        })

final_rows.extend(viat_rows)
print(f"Viáticos rows added: {len(viat_rows)} locales")


# ── 6. Per-local summary ──────────────────────────────────────────────────────

local_sum = {}
for r in final_rows:
    k = r["CÓDIGO"]
    if k not in local_sum:
        local_sum[k] = {
            "CÓDIGO":          k,
            "CC_ORIGINAL":     r["CC_ORIGINAL"],
            "MARCA":           r["MARCA"],
            "NOMBRE_LOCAL":    r["NOMBRE_LOCAL"],
            "MES_SERVICIO":    r["MES_SERVICIO"],
            "N_EXTINTORES":    0,
            "TOTAL_MANTT":     0.0,
            "TOTAL_RECARGA":   0.0,
            "COBRO_ANUAL":     0.0,
            "AÑO_RECARGA":     r["AÑO_RECARGA"],
            "AÑO_ULT_RECARGA": r["AÑO_ULT_RECARGA"],
            "_desglose":       [],
        }
    s = local_sum[k]
    if r["TIPO"] != "MOVILIZACIÓN":
        s["N_EXTINTORES"] += 1
    # Update AÑO_RECARGA from first non-movilización row
    if r["TIPO"] != "MOVILIZACIÓN" and r["AÑO_RECARGA"] is not None:
        s["AÑO_RECARGA"]     = r["AÑO_RECARGA"]
        s["AÑO_ULT_RECARGA"] = r["AÑO_ULT_RECARGA"]
    s["TOTAL_MANTT"]   += r["_cm"]
    s["TOTAL_RECARGA"] += r["_cr"]
    s["COBRO_ANUAL"]   += r["_cobro"]
    s["_desglose"].append(
        f"{r['TIPO']} {r['CAPACIDAD']}=${r['_cobro']:.2f}"
    )

for s in local_sum.values():
    s["TOTAL_MANTT"]  = round(s["TOTAL_MANTT"],  2)
    s["TOTAL_RECARGA"]= round(s["TOTAL_RECARGA"], 2)
    s["COBRO_ANUAL"]  = round(s["COBRO_ANUAL"],  2)
    s["DESGLOSE"]     = " | ".join(s.pop("_desglose"))

n_locales   = len(local_sum)
total_anual = sum(s["COBRO_ANUAL"] for s in local_sum.values())

print(f"\n{'='*50}")
print(f"LOCALES TOTAL: {n_locales}  (target: 198)")
print(f"COBRO ANUAL TOTAL: ${total_anual:,.2f}")
print(f"{'='*50}")

print("\nLocales por mes:")
mc = Counter(s["MES_SERVICIO"] for s in local_sum.values())
for m in MONTHS:
    if mc.get(m): print(f"  {m:15s}: {mc[m]}")

print("\nLocales por marca:")
bc = Counter(s["MARCA"] for s in local_sum.values())
for b, c in sorted(bc.items(), key=lambda x: -x[1]):
    print(f"  {b:25s}: {c}")

# Audit: in PRESUPUESTO (minus G042, H068) but not in final.
# Las CN compartidas (CN04/CN16/CN31/CN37) están cubiertas por su BS equivalente.
SHARED_CN = set(SHARED_BS_CN.values())   # {('CN',4),('CN',16),('CN',31),('CN',37)}
pres_expected = set(
    k for k in presupuesto_map
    if k not in (G042_KEY, H068_KEY) and k not in SHARED_CN and k not in DELETE_LOCALS
)
final_codes   = set(r["_ckey"] for r in final_rows)
missing_from_final = pres_expected - final_codes
extra_in_final     = final_codes - pres_expected - {H068_KEY}

print(f"\nIn PRESUPUESTO but missing from DB ({len(missing_from_final)}):")
for k in sorted(missing_from_final):
    info = presupuesto_map[k]
    print(f"  {fmt_code(*k)}: {info['cc_pres']} – {info['marca']} – {info['nombre']}")

print(f"\nIn DB but not in PRESUPUESTO ({len(extra_in_final)}):")
for k in sorted(extra_in_final):
    print(f"  {fmt_code(*k)}")


# ── 7. Write Excel ────────────────────────────────────────────────────────────

OUT = "/home/user/previfuego-facturacion/BASE_DATOS_KFC.xlsx"
wb = openpyxl.Workbook()

from openpyxl.styles import Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Styles ──
COL_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue – column headers
LOCAL_FILL    = PatternFill("solid", fgColor="D9E1F2")   # light blue – each local's rows
TOTALS_FILL   = PatternFill("solid", fgColor="E2EFDA")   # light green – TOTALES row
RESUMEN_FILL  = PatternFill("solid", fgColor="375623")   # dark green – resumen header
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT      = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BOLD_FONT     = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT   = Font(name="Calibri", size=11)
TOTALS_FONT   = Font(name="Calibri", bold=True, size=11, color="375623")

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
RIGHT   = Alignment(horizontal="right",  vertical="center")

CURRENCY_FMT = '$#,##0.00'
THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

month_sort = {m: i for i, m in enumerate(MONTHS)}

# Group extintores by (mes, codigo) in sorted order
from collections import OrderedDict

sorted_rows = sorted(
    final_rows,
    key=lambda x: (month_sort.get(x["MES_SERVICIO"], 99), x["CÓDIGO"])
)

# Build ordered list of locals preserving sort
local_order = list(OrderedDict.fromkeys(r["CÓDIGO"] for r in sorted_rows))

# Map código → list of its rows (in order)
rows_by_local = OrderedDict()
for code in local_order:
    rows_by_local[code] = [r for r in sorted_rows if r["CÓDIGO"] == code]


# ── Sheet 1: DETALLE ─────────────────────────────────────────────────────────
ws_det = wb.active
ws_det.title = "DETALLE"
ws_det.sheet_view.showGridLines = False
ws_det.row_dimensions[1].height = 22

COLS_DET = [
    "MARCA", "NOMBRE_LOCAL", "MES_SERVICIO", "UBICACIÓN",
    "TIPO", "CAPACIDAD", "COSTO_MANTT", "PRECIO_RECARGA",
    "AÑO_ULT_RECARGA", "AÑO_RECARGA"
]

# Fixed column widths (chars) for readability
COL_WIDTHS = [28, 38, 15, 30, 10, 12, 16, 18, 18, 14]

# Write header row
for ci, (col_name, width) in enumerate(zip(COLS_DET, COL_WIDTHS), 1):
    cell = ws_det.cell(row=1, column=ci, value=col_name)
    cell.fill = COL_HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws_det.column_dimensions[get_column_letter(ci)].width = width

cur_row = 2
price_cols = {COLS_DET.index("COSTO_MANTT") + 1, COLS_DET.index("PRECIO_RECARGA") + 1}

for code in local_order:
    group = rows_by_local[code]
    first = group[0]

    # Write each extintor row (and movilización if present)
    ext_group    = [r for r in group if r["TIPO"] != "MOVILIZACIÓN"]
    mov_group    = [r for r in group if r["TIPO"] == "MOVILIZACIÓN"]

    MOVIL_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow for viáticos

    for r in ext_group:
        ws_det.row_dimensions[cur_row].height = 18
        values = [
            r["MARCA"], r["NOMBRE_LOCAL"], r["MES_SERVICIO"], r["UBICACIÓN"],
            r["TIPO"], r["CAPACIDAD"],
            r["COSTO_MANTT"], r["PRECIO_RECARGA"],
            r["AÑO_ULT_RECARGA"], r["AÑO_RECARGA"],
        ]
        for ci, val in enumerate(values, 1):
            cell = ws_det.cell(row=cur_row, column=ci, value=val)
            cell.fill = LOCAL_FILL
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if ci in price_cols:
                cell.number_format = CURRENCY_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
        cur_row += 1

    for r in mov_group:
        ws_det.row_dimensions[cur_row].height = 18
        values = [
            r["MARCA"], r["NOMBRE_LOCAL"], r["MES_SERVICIO"], r["UBICACIÓN"],
            r["TIPO"], r["CAPACIDAD"],
            r["COSTO_MANTT"], r["PRECIO_RECARGA"],
            r["AÑO_ULT_RECARGA"], r["AÑO_RECARGA"],
        ]
        for ci, val in enumerate(values, 1):
            cell = ws_det.cell(row=cur_row, column=ci, value=val)
            cell.fill = MOVIL_FILL
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if ci in price_cols:
                cell.number_format = CURRENCY_FMT
                cell.alignment = RIGHT
            else:
                cell.alignment = LEFT
        cur_row += 1

    # TOTALES row (includes movilización)
    total_mantt   = round(sum(r["_cm"] for r in group), 2)
    total_recarga = round(sum(r["_cr"] for r in group), 2)
    ws_det.row_dimensions[cur_row].height = 20
    n_ext_only = len(ext_group)
    mov_note   = f" + movil.${mov_group[0]['_cm']:.2f}" if mov_group else ""
    totals_values = [
        f"TOTALES  {first['MARCA']} – {first['NOMBRE_LOCAL']}",
        "", "", "",
        "", f"{n_ext_only} ext.{mov_note}",
        total_mantt, total_recarga,
        "", "",
    ]
    for ci, val in enumerate(totals_values, 1):
        cell = ws_det.cell(row=cur_row, column=ci, value=val)
        cell.fill = TOTALS_FILL
        cell.font = TOTALS_FONT
        cell.border = THIN_BORDER
        if ci in price_cols:
            cell.number_format = CURRENCY_FMT
            cell.alignment = RIGHT
        else:
            cell.alignment = LEFT
    # Merge first 4 cells of TOTALES label
    ws_det.merge_cells(
        start_row=cur_row, start_column=1,
        end_row=cur_row, end_column=4
    )
    cur_row += 1

    # Blank separator row
    ws_det.row_dimensions[cur_row].height = 8
    cur_row += 1


# ── Sheet 2: RESUMEN_LOCALES ─────────────────────────────────────────────────
ws_res = wb.create_sheet("RESUMEN_LOCALES")
ws_res.sheet_view.showGridLines = False
ws_res.row_dimensions[1].height = 22

COLS_RES = [
    "CÓDIGO", "MARCA", "NOMBRE_LOCAL", "MES_SERVICIO",
    "N_EXTINTORES", "TOTAL_MANTT ($)", "TOTAL_RECARGA ($)",
    "COBRO_ANUAL ($)", "AÑO_ULT_RECARGA", "AÑO_RECARGA"
]
RES_WIDTHS = [10, 28, 38, 15, 14, 18, 18, 18, 18, 14]
RES_PRICE_COLS = {6, 7, 8}  # 1-indexed columns for currency

for ci, (col_name, width) in enumerate(zip(COLS_RES, RES_WIDTHS), 1):
    cell = ws_res.cell(row=1, column=ci, value=col_name)
    cell.fill = RESUMEN_FILL
    cell.font = HDR_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
    ws_res.column_dimensions[get_column_letter(ci)].width = width

ALT_FILLS = [
    PatternFill("solid", fgColor="EBF3FB"),
    PatternFill("solid", fgColor="FFFFFF"),
]

for ridx, code in enumerate(local_order, 2):
    s = local_sum[code]
    ws_res.row_dimensions[ridx].height = 18
    values = [
        s["CÓDIGO"], s["MARCA"], s["NOMBRE_LOCAL"], s["MES_SERVICIO"],
        s["N_EXTINTORES"],
        s["TOTAL_MANTT"], s["TOTAL_RECARGA"], s["COBRO_ANUAL"],
        s["AÑO_ULT_RECARGA"], s["AÑO_RECARGA"],
    ]
    row_fill = ALT_FILLS[(ridx) % 2]
    for ci, val in enumerate(values, 1):
        cell = ws_res.cell(row=ridx, column=ci, value=val)
        cell.fill = row_fill
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if ci in RES_PRICE_COLS:
            cell.number_format = CURRENCY_FMT
            cell.alignment = RIGHT
        else:
            cell.alignment = LEFT

# Grand total row
grand_row = len(local_order) + 2
ws_res.row_dimensions[grand_row].height = 22
grand_mantt   = round(sum(local_sum[c]["TOTAL_MANTT"]   for c in local_order), 2)
grand_recarga = round(sum(local_sum[c]["TOTAL_RECARGA"] for c in local_order), 2)
grand_cobro   = round(sum(local_sum[c]["COBRO_ANUAL"]   for c in local_order), 2)
grand_vals = [
    "GRAN TOTAL", "", f"{len(local_order)} LOCALES", "",
    sum(local_sum[c]["N_EXTINTORES"] for c in local_order),
    grand_mantt, grand_recarga, grand_cobro, "", "",
]
for ci, val in enumerate(grand_vals, 1):
    cell = ws_res.cell(row=grand_row, column=ci, value=val)
    cell.fill = COL_HDR_FILL
    cell.font = HDR_FONT
    cell.border = THIN_BORDER
    if ci in RES_PRICE_COLS:
        cell.number_format = CURRENCY_FMT
        cell.alignment = RIGHT
    else:
        cell.alignment = CENTER


wb.save(OUT)
print(f"\n✓ Saved: {OUT}")
print(f"  DETALLE rows:  {len(final_rows)}")
print(f"  Locales:       {n_locales}")
print(f"  Cobro anual:   ${total_anual:,.2f}")
