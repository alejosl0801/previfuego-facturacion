#!/usr/bin/env python3
"""
Reconstruye el DETALLE de los locales SUSHICORP (KN001-KN013) en
BASE_DATOS_KFC.xlsx a partir de BASE_DATOS_SUSHICORP.xlsx, que contiene
el detalle completo de los 62 extintores.

Verifica que la suma de costos por local coincida EXACTAMENTE con el
RESUMEN_LOCALES antes de escribir nada. Si algún local no cuadra, aborta.
"""

import openpyxl
import re
import sys

KFC = "BASE_DATOS_KFC.xlsx"
SUSHI = "BASE_DATOS_SUSHICORP.xlsx"

# Normaliza tipo SUSHICORP -> formato KFC
TIPO_MAP = {"K": "TIPO K"}


def tipo_fmt(t):
    t = str(t or "").strip().upper()
    return TIPO_MAP.get(t, t)


def build_precio_recarga_map():
    """(tipo, cap, costo_mantt) -> precio_recarga, desde el DETALLE del KFC."""
    wb = openpyxl.load_workbook(KFC, read_only=True)
    mapa = {}
    for r in list(wb["DETALLE"].iter_rows(values_only=True))[1:]:
        tipo = str(r[4]).strip().upper() if r[4] else ""
        cap = str(r[5]).strip().upper() if r[5] else ""
        cm, pr = r[6], r[7]
        if tipo and cap and cm is not None and pr is not None:
            mapa[(tipo, cap, float(cm))] = float(pr)
    wb.close()
    return mapa


def norm_name(s):
    """Quita prefijo 'KNxx - ' y normaliza para comparar nombres."""
    s = str(s or "").strip().upper()
    s = re.sub(r"^KN\d+\s*-\s*", "", s)
    return s


def cap_fmt(num, tipo):
    """50 -> '50 LBS'; tipo K usa galones: '2,5G' -> '2.5 GLS'."""
    s = str(num or "").strip().upper().replace(",", ".")
    if tipo_fmt(tipo) == "TIPO K":
        m = re.search(r"([\d.]+)", s)
        return f"{m.group(1)} GLS" if m else s
    try:
        n = int(float(s))
        return f"{n} LBS"
    except (TypeError, ValueError):
        return s


def main():
    PRECIO_REC = build_precio_recarga_map()
    wb_k = openpyxl.load_workbook(KFC)
    res = list(wb_k["RESUMEN_LOCALES"].iter_rows(values_only=True))

    # Mapa nombre_normalizado -> fila resumen (codigo, nombre_local, mes, tMantt, tRec)
    resumen_kn = {}
    for r in res[1:]:
        if r[0] and str(r[0]).upper().startswith("KN"):
            resumen_kn[norm_name(r[2])] = {
                "codigo": r[0],
                "nombre_local": r[2],
                "mes": str(r[3]).strip().upper(),
                "n_ext": r[4],
                "t_mantt": float(r[5]),
                "t_rec": float(r[6]),
                "ult_rec": r[8],
                "anio_rec": r[9],
            }

    # Leer SUSHICORP y agrupar extintores por local
    wb_s = openpyxl.load_workbook(SUSHI, read_only=True)
    rows = list(wb_s["SUSHICORP"].iter_rows(values_only=True))

    grupos = []
    actual = None
    for r in rows[1:]:
        cc = r[0]
        tipo = r[2]
        if cc and "@" not in str(cc):  # nuevo local (ignora emails)
            actual = {"nombre": str(cc).strip(), "exts": []}
            grupos.append(actual)
        if actual and tipo:
            actual["exts"].append(
                {
                    "ubicacion": str(r[1] or "").strip(),
                    "tipo": tipo_fmt(r[2]),
                    "cap": cap_fmt(r[3], r[2]),
                    "costo_mantt": float(r[7]) if r[7] is not None else None,
                }
            )

    # Construir filas detalle nuevas + verificar contra resumen
    nuevas_filas = []
    errores = []
    for g in grupos:
        key = norm_name(g["nombre"])
        info = resumen_kn.get(key)
        if not info:
            errores.append(f"Local SUSHICORP '{g['nombre']}' no encontrado en RESUMEN_LOCALES")
            continue

        suma_mantt = 0.0
        suma_rec = 0.0
        for e in g["exts"]:
            cm = e["costo_mantt"]
            pk = (e["tipo"], e["cap"], cm)
            if cm is None or pk not in PRECIO_REC:
                errores.append(f"{info['codigo']}: sin precio_recarga para {pk}")
                continue
            pr = PRECIO_REC[pk]
            suma_mantt += cm
            suma_rec += pr
            nuevas_filas.append([
                "SUSHICORP",
                info["nombre_local"],
                info["mes"],
                e["ubicacion"],
                e["tipo"],
                e["cap"],
                cm,
                pr,
                info["ult_rec"],
                info["anio_rec"],
            ])

        # Verificacion estricta
        if len(g["exts"]) != info["n_ext"]:
            errores.append(f"{info['codigo']}: {len(g['exts'])} extintores vs N_EXTINTORES={info['n_ext']}")
        if round(suma_mantt, 2) != round(info["t_mantt"], 2):
            errores.append(f"{info['codigo']}: mantt {suma_mantt} vs resumen {info['t_mantt']}")
        if round(suma_rec, 2) != round(info["t_rec"], 2):
            errores.append(f"{info['codigo']}: recarga {suma_rec} vs resumen {info['t_rec']}")

    if errores:
        print("ABORTADO — inconsistencias detectadas:")
        for e in errores:
            print("  -", e)
        sys.exit(1)

    print(f"Verificacion OK: {len(nuevas_filas)} extintores reconstruidos, todos cuadran con el resumen.")

    # Reescribir DETALLE: conservar filas no-KN, reemplazar las KN
    ws_det = wb_k["DETALLE"]
    det_rows = list(ws_det.iter_rows(values_only=True))
    header = det_rows[0]
    conservadas = [
        list(r) for r in det_rows[1:]
        if not (r[1] and str(r[1]).upper().startswith("KN"))
    ]

    # Limpiar hoja y reescribir
    wb_k.remove(ws_det)
    ws_new = wb_k.create_sheet("DETALLE")
    # Mover DETALLE a la posición 2 (después de RESUMEN_LOCALES)
    wb_k.move_sheet("DETALLE", -(len(wb_k.sheetnames) - 2))
    ws_new.append(list(header))
    for r in conservadas:
        ws_new.append(r)
    for r in nuevas_filas:
        ws_new.append(r)

    wb_k.save(KFC)
    total = len(conservadas) + len(nuevas_filas)
    print(f"DETALLE reescrito: {len(conservadas)} filas existentes + {len(nuevas_filas)} KN = {total} extintores totales.")


if __name__ == "__main__":
    main()
