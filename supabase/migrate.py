#!/usr/bin/env python3
"""
Migra los datos del Excel (BASE_DATOS_KFC.xlsx) a Supabase.

Uso:
  1. Instala dependencias:  pip install openpyxl requests
  2. Configura variables de entorno:
       export SUPABASE_URL="https://xxxxx.supabase.co"
       export SUPABASE_SERVICE_KEY="eyJhbGciOi..."   # service_role key (NO la anon key)
  3. Ejecuta:  python migrate.py

El script lee las hojas RESUMEN_LOCALES y DETALLE, normaliza los códigos
de local, y sube todo a las tablas `locales` y `extintores` en Supabase.
"""

import os
import re
import sys
import json
import time

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl:  pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("Instala requests:  pip install requests")
    sys.exit(1)

SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: configura SUPABASE_URL y SUPABASE_SERVICE_KEY como variables de entorno.")
    sys.exit(1)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

CN_TO_BS = {"CN004": "BS004", "CN016": "BS016", "CN031": "BS031", "CN037": "BS037"}


def norm_code(raw):
    m = re.match(r"^([A-Z]+)(\d+)", str(raw).strip().upper())
    if not m:
        return str(raw).strip().upper()
    prefix, num = m.group(1), m.group(2)
    code = prefix + str(int(num)).zfill(3)
    return CN_TO_BS.get(code, code)


def safe_float(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def safe_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def api_delete(table, params):
    r = requests.delete(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS, params=params)
    if r.status_code not in (200, 204):
        print(f"  DELETE {table}: {r.status_code} — {r.text[:200]}")
    return r


def api_post(table, data, batch_size=200):
    total = 0
    for i in range(0, len(data), batch_size):
        batch = data[i : i + batch_size]
        r = requests.post(f"{SUPABASE_URL}/rest/v1/{table}", headers=HEADERS, json=batch)
        if r.status_code not in (200, 201):
            print(f"  POST {table} batch {i // batch_size + 1}: {r.status_code} — {r.text[:300]}")
            return total
        total += len(batch)
    return total


def migrate(excel_path):
    print(f"Leyendo {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ── RESUMEN_LOCALES ───────────────────────────────────────────────────────
    ws_res = wb["RESUMEN_LOCALES"]
    locales = []
    seen_codes = set()

    for row in ws_res.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        codigo = norm_code(row[0])
        if codigo in seen_codes:
            print(f"  Duplicado en RESUMEN_LOCALES: {row[0]} → {codigo} (ignorado)")
            continue
        seen_codes.add(codigo)

        locales.append(
            {
                "codigo": codigo,
                "marca": str(row[1] or "").strip(),
                "nombre_local": str(row[2] or codigo).strip(),
                "mes_servicio": str(row[3] or "").strip().upper(),
                "n_extintores": safe_int(row[4]) or 0,
                "total_mantt": safe_float(row[5]),
                "total_recarga": safe_float(row[6]),
                "cobro_anual": safe_float(row[7]),
                "anio_ult_recarga": safe_int(row[8]),
                "anio_recarga": safe_int(row[9]),
            }
        )

    print(f"  {len(locales)} locales leídos del Excel")

    # ── DETALLE ───────────────────────────────────────────────────────────────
    ws_det = wb["DETALLE"]
    extintores = []

    for row in ws_det.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        m = re.match(r"^([A-Z]+\d+)", str(row[1]).strip().upper())
        if not m:
            continue
        codigo_local = norm_code(m.group(1))
        if codigo_local not in seen_codes:
            continue

        extintores.append(
            {
                "codigo_local": codigo_local,
                "marca": str(row[0] or "").strip(),
                "nombre_local": str(row[1] or "").strip(),
                "mes_servicio": str(row[2] or "").strip().upper(),
                "ubicacion": str(row[3] or "").strip(),
                "tipo": str(row[4] or "").strip().upper(),
                "capacidad": str(row[5] or "").strip(),
                "costo_mantt": safe_float(row[6]),
                "precio_recarga": safe_float(row[7]),
                "anio_ult_recarga": safe_int(row[8]),
                "anio_recarga": safe_int(row[9]),
            }
        )

    print(f"  {len(extintores)} extintores leídos del Excel")

    # ── SUBIR A SUPABASE ──────────────────────────────────────────────────────
    print("\nLimpiando tablas en Supabase...")
    api_delete("extintores", {"id": "gt.0"})
    api_delete("locales", {"codigo": "not.is.null"})
    time.sleep(0.5)

    print("Subiendo locales...")
    n = api_post("locales", locales)
    print(f"  {n}/{len(locales)} locales subidos")

    print("Subiendo extintores...")
    n = api_post("extintores", extintores)
    print(f"  {n}/{len(extintores)} extintores subidos")

    print("\nMigración completada.")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "BASE_DATOS_KFC.xlsx"
    if not os.path.exists(path):
        print(f"Archivo no encontrado: {path}")
        sys.exit(1)
    migrate(path)
