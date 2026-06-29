#!/usr/bin/env python3
"""
Separa los datos en DOS bases independientes:

  BASE_DATOS_KFC.xlsx       -> solo grupo KFC (197 locales, SIN los KN)
  BASE_DATOS_SUSHICORP.xlsx -> solo SUSHICORP/KOBE/NOE (13 locales, 62 extintores)

Fuentes:
  - KFC_SRC:   el Excel KFC subido (tiene 210 locales, incluidos 13 KN a extraer)
  - SUSHI_SRC: el Excel SUSHICORP subido (hoja matriz con el detalle real)

Los totales de los locales SUSHICORP se toman tal cual del RESUMEN del KFC
(ya validados), y el detalle de extintores se reconstruye de la matriz SUSHICORP,
verificando que las sumas cuadren EXACTO antes de escribir.
"""

import openpyxl
import re
import sys

KFC_SRC = "/root/.claude/uploads/1ed2eba0-307e-5f58-bac6-767662837fdb/2277c709-BASE_DATOS_KFC.xlsx"
SUSHI_SRC = "/root/.claude/uploads/1ed2eba0-307e-5f58-bac6-767662837fdb/a7c9a1c7-BASE_DATOS_SUSHICORP.xlsx"

KFC_OUT = "BASE_DATOS_KFC.xlsx"
SUSHI_OUT = "BASE_DATOS_SUSHICORP.xlsx"

TIPO_MAP = {"K": "TIPO K"}


def tipo_fmt(t):
    return TIPO_MAP.get(str(t or "").strip().upper(), str(t or "").strip().upper())


def cap_fmt(num, tipo):
    s = str(num or "").strip().upper().replace(",", ".")
    if tipo_fmt(tipo) == "TIPO K":
        m = re.search(r"([\d.]+)", s)
        return f"{m.group(1)} GLS" if m else s
    try:
        return f"{int(float(s))} LBS"
    except (TypeError, ValueError):
        return s


def norm_name(s):
    s = str(s or "").strip().upper()
    return re.sub(r"^KN\d+\s*-\s*", "", s)


def build_precio_recarga_map(det_rows):
    """(tipo, cap, costo_mantt) -> precio_recarga, desde el DETALLE del KFC (sin KN)."""
    mapa = {}
    for r in det_rows[1:]:
        tipo = str(r[4]).strip().upper() if r[4] else ""
        cap = str(r[5]).strip().upper() if r[5] else ""
        cm, pr = r[6], r[7]
        if tipo and cap and cm is not None and pr is not None and "ext" not in cap.lower():
            mapa[(tipo, cap, float(cm))] = float(pr)
    return mapa


def main():
    wb_k = openpyxl.load_workbook(KFC_SRC)
    res_rows = list(wb_k["RESUMEN_LOCALES"].iter_rows(values_only=True))
    det_rows = list(wb_k["DETALLE"].iter_rows(values_only=True))
    res_header = res_rows[0]
    det_header = det_rows[0]

    # Capturar info de los 13 KN del RESUMEN del KFC (fuente de totales)
    kn_resumen = {}
    for r in res_rows[1:]:
        if r[0] and str(r[0]).upper().startswith("KN"):
            kn_resumen[norm_name(r[2])] = {
                "row": list(r),
                "nombre": r[2],
                "mes": str(r[3]).strip().upper(),
                "n_ext": int(r[4]) if r[4] else 0,
                "t_mantt": round(float(r[5]), 2) if r[5] else 0,
                "t_rec": round(float(r[6]), 2) if r[6] else 0,
                "ult_rec": r[8],
                "anio_rec": r[9],
            }

    PRECIO_REC = build_precio_recarga_map(det_rows)

    # ── 1. Construir DETALLE de SUSHICORP desde la matriz ────────────────────
    wb_s = openpyxl.load_workbook(SUSHI_SRC, read_only=True)
    rows = list(wb_s["SUSHICORP"].iter_rows(values_only=True))
    grupos = []
    actual = None
    for r in rows[1:]:
        cc = r[0]
        if cc and "@" not in str(cc):
            actual = {"nombre": str(cc).strip(), "exts": []}
            grupos.append(actual)
        if actual and r[2]:
            actual["exts"].append({
                "ubic": str(r[1] or "").strip(),
                "tipo": tipo_fmt(r[2]),
                "cap": cap_fmt(r[3], r[2]),
                "cm": float(r[7]) if r[7] is not None else None,
            })

    sushi_detalle = []
    errores = []
    for g in grupos:
        info = kn_resumen.get(norm_name(g["nombre"]))
        if not info:
            errores.append(f"Local SUSHICORP '{g['nombre']}' no está en el RESUMEN del KFC")
            continue
        sm = sr = 0.0
        for e in g["exts"]:
            pk = (e["tipo"], e["cap"], e["cm"])
            if e["cm"] is None or pk not in PRECIO_REC:
                errores.append(f"{info['nombre']}: sin precio_recarga para {pk}")
                continue
            pr = PRECIO_REC[pk]
            sm += e["cm"]
            sr += pr
            sushi_detalle.append([
                "SUSHICORP", info["nombre"], info["mes"], e["ubic"],
                e["tipo"], e["cap"], e["cm"], pr, info["ult_rec"], info["anio_rec"],
            ])
        if len(g["exts"]) != info["n_ext"]:
            errores.append(f"{info['nombre']}: {len(g['exts'])} ext vs resumen {info['n_ext']}")
        if round(sm, 2) != info["t_mantt"]:
            errores.append(f"{info['nombre']}: mantt {round(sm,2)} vs resumen {info['t_mantt']}")
        if round(sr, 2) != info["t_rec"]:
            errores.append(f"{info['nombre']}: recarga {round(sr,2)} vs resumen {info['t_rec']}")

    if errores:
        print("ABORTADO — inconsistencias:")
        for e in errores:
            print("  -", e)
        sys.exit(1)

    print(f"SUSHICORP: {len(grupos)} locales, {len(sushi_detalle)} extintores — verificado contra resumen KFC.")

    # ── 2. Escribir BASE_DATOS_SUSHICORP.xlsx ────────────────────────────────
    wbS = openpyxl.Workbook()
    wsR = wbS.active
    wsR.title = "RESUMEN_LOCALES"
    wsR.append(list(res_header))
    for info in sorted(kn_resumen.values(), key=lambda x: x["row"][0]):
        wsR.append(info["row"])
    wsD = wbS.create_sheet("DETALLE")
    wsD.append(list(det_header))
    for r in sushi_detalle:
        wsD.append(r)
    wbS.save(SUSHI_OUT)
    print(f"  -> {SUSHI_OUT} escrito ({len(kn_resumen)} locales, {len(sushi_detalle)} extintores).")

    # ── 3. Escribir BASE_DATOS_KFC.xlsx SIN los KN ───────────────────────────
    # RESUMEN sin KN
    ws_res = wb_k["RESUMEN_LOCALES"]
    keep_res = [list(r) for r in res_rows[1:] if not (r[0] and str(r[0]).upper().startswith("KN"))]
    wb_k.remove(ws_res)
    ws_res2 = wb_k.create_sheet("RESUMEN_LOCALES")
    wb_k.move_sheet("RESUMEN_LOCALES", -(len(wb_k.sheetnames) - 1))
    ws_res2.append(list(res_header))
    for r in keep_res:
        ws_res2.append(r)

    # DETALLE sin KN (filas cuyo NOMBRE_LOCAL empieza con KN, con forward-fill de bloque)
    ws_det = wb_k["DETALLE"]
    keep_det = []
    es_kn_actual = False
    for r in det_rows[1:]:
        cap = str(r[5] or "").lower()
        if "ext" in cap:  # subtotal: pertenece al bloque actual
            if not es_kn_actual:
                keep_det.append(list(r))
            continue
        if r[1] and re.match(r"^[A-Z]+\d+", str(r[1]).strip().upper()):
            es_kn_actual = str(r[1]).strip().upper().startswith("KN")
        if not any(c is not None for c in r):
            es_kn_actual = False
            keep_det.append(list(r))
            continue
        if not es_kn_actual:
            keep_det.append(list(r))
    wb_k.remove(ws_det)
    ws_det2 = wb_k.create_sheet("DETALLE")
    wb_k.move_sheet("DETALLE", -(len(wb_k.sheetnames) - 2))
    ws_det2.append(list(det_header))
    for r in keep_det:
        ws_det2.append(r)

    wb_k.save(KFC_OUT)
    print(f"  -> {KFC_OUT} escrito ({len(keep_res)} locales, sin SUSHICORP).")


if __name__ == "__main__":
    main()
