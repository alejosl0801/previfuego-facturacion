"""
Script para separar los locales SUSHICORP (prefijo KN) de BASE_DATOS_KFC.xlsx
a un nuevo archivo BASE_DATOS_SUSHICORP.xlsx.

- BASE_DATOS_KFC.xlsx: conserva solo filas no-KN, GRAN TOTAL actualizado
- BASE_DATOS_SUSHICORP.xlsx: contiene solo filas KN, GRAN TOTAL de KN
"""

import openpyxl
from openpyxl import load_workbook
from copy import copy
import os

SRC = "/home/user/previfuego-facturacion/BASE_DATOS_KFC.xlsx"
DST_SUSHI = "/home/user/previfuego-facturacion/BASE_DATOS_SUSHICORP.xlsx"


def copy_cell(src_cell, dst_cell):
    """Copia valor y formato básico de una celda a otra."""
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def copy_row(src_ws, src_row_idx, dst_ws, dst_row_idx):
    """Copia una fila completa (valores + formato) entre worksheets."""
    for col_idx in range(1, src_ws.max_column + 1):
        src_cell = src_ws.cell(row=src_row_idx, column=col_idx)
        dst_cell = dst_ws.cell(row=dst_row_idx, column=col_idx)
        copy_cell(src_cell, dst_cell)


def copy_column_dimensions(src_ws, dst_ws):
    """Copia ancho de columnas."""
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width


def identify_kn_rows_detalle(ws):
    """
    Retorna un set de índices de fila (1-based) en DETALLE que pertenecen
    a bloques KN (SUSHICORP).

    Estructura de bloque:
      - Primera fila: col A = 'SUSHICORP', col B starts with 'KN'
      - Filas internas: col A = None (datos del extintor)
      - Fila TOTALES: col A starts with 'TOTALES  SUSHICORP'
      - Fila blank separadora (opcional): todos None
    """
    rows = list(ws.iter_rows(values_only=True))
    kn_indices = set()

    i = 0
    while i < len(rows):
        row = rows[i]
        col_a = str(row[0]) if row[0] else ''
        col_b = str(row[1]) if row[1] else ''

        if col_a == 'SUSHICORP' and col_b.startswith('KN'):
            kn_indices.add(i + 1)  # 1-based
            i += 1
            # Collect inner rows until TOTALES
            while i < len(rows):
                r = rows[i]
                kn_indices.add(i + 1)
                col_a2 = str(r[0]) if r[0] else ''
                if col_a2.startswith('TOTALES') and 'SUSHICORP' in col_a2:
                    break
                i += 1
            i += 1
            # Include trailing blank separator row
            if i < len(rows) and all(v is None for v in rows[i]):
                kn_indices.add(i + 1)
        i += 1

    return kn_indices


def compute_gran_total_resumen(ws):
    """
    Calcula los valores para la fila GRAN TOTAL de RESUMEN_LOCALES.
    Suma col E (n_ext), F (total_mantt), G (total_recarga), H (cobro_anual).
    Retorna (count_locales, sum_ext, sum_mantt, sum_recarga, sum_cobro).
    """
    count = 0
    sum_ext = 0
    sum_mantt = 0.0
    sum_recarga = 0.0
    sum_cobro = 0.0

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0]) if row[0] else ''
        if not code or code == 'GRAN TOTAL':
            continue
        count += 1
        sum_ext += (row[4] or 0)
        sum_mantt += (row[5] or 0)
        sum_recarga += (row[6] or 0)
        sum_cobro += (row[7] or 0)

    # Round to avoid floating point noise
    return (count,
            sum_ext,
            round(sum_mantt, 2),
            round(sum_recarga, 2),
            round(sum_cobro, 2))


def process_resumen_locales(src_wb, dst_sushi_wb):
    """
    Procesa la hoja RESUMEN_LOCALES:
    - Copia filas KN a dst_sushi_wb
    - Elimina filas KN de src_wb
    - Actualiza GRAN TOTAL en ambos workbooks
    """
    src_ws = src_wb['RESUMEN_LOCALES']

    # Create destination sheet preserving order as in original
    sushi_ws = dst_sushi_wb.create_sheet('RESUMEN_LOCALES')
    copy_column_dimensions(src_ws, sushi_ws)

    # Collect all rows
    all_rows = list(src_ws.iter_rows())  # cell objects (not values_only)

    # Separate header, data rows (KN vs non-KN), and GRAN TOTAL
    header_row = all_rows[0]
    gran_total_row = None
    kn_data_rows = []
    non_kn_data_rows = []

    for row in all_rows[1:]:
        code = str(row[0].value) if row[0].value else ''
        if code == 'GRAN TOTAL':
            gran_total_row = row
        elif code.startswith('KN'):
            kn_data_rows.append(row)
        else:
            non_kn_data_rows.append(row)

    print(f"RESUMEN_LOCALES: {len(kn_data_rows)} filas KN, {len(non_kn_data_rows)} filas no-KN")

    # --- Build SUSHICORP sheet ---
    dst_row = 1
    # Header
    for col_idx, cell in enumerate(header_row, 1):
        copy_cell(cell, sushi_ws.cell(row=dst_row, column=col_idx))
    dst_row += 1
    # KN data
    for src_row in kn_data_rows:
        for col_idx, cell in enumerate(src_row, 1):
            copy_cell(cell, sushi_ws.cell(row=dst_row, column=col_idx))
        dst_row += 1
    # GRAN TOTAL placeholder (will be updated below)
    if gran_total_row:
        for col_idx, cell in enumerate(gran_total_row, 1):
            copy_cell(cell, sushi_ws.cell(row=dst_row, column=col_idx))

    # --- Rebuild KFC sheet in-place ---
    # We'll delete all data rows then rewrite them
    # Delete rows from bottom to top (to preserve indices)
    rows_to_delete = []
    for row in src_ws.iter_rows(min_row=2):
        code = str(row[0].value) if row[0].value else ''
        if code.startswith('KN'):
            rows_to_delete.append(row[0].row)

    # Delete from bottom to top
    for row_num in sorted(rows_to_delete, reverse=True):
        src_ws.delete_rows(row_num)

    # Update GRAN TOTAL in KFC
    (count_kfc, ext_kfc, mantt_kfc, rec_kfc, cobro_kfc) = compute_gran_total_resumen(src_ws)
    gt_row_kfc = src_ws.max_row  # GRAN TOTAL is the last row
    src_ws.cell(row=gt_row_kfc, column=3).value = f"{count_kfc} LOCALES"
    src_ws.cell(row=gt_row_kfc, column=5).value = ext_kfc
    src_ws.cell(row=gt_row_kfc, column=6).value = mantt_kfc
    src_ws.cell(row=gt_row_kfc, column=7).value = rec_kfc
    src_ws.cell(row=gt_row_kfc, column=8).value = cobro_kfc
    print(f"KFC GRAN TOTAL actualizado: {count_kfc} locales, {ext_kfc} ext, mantt={mantt_kfc}, rec={rec_kfc}, cobro={cobro_kfc}")

    # Update GRAN TOTAL in SUSHICORP sheet
    (count_sushi, ext_sushi, mantt_sushi, rec_sushi, cobro_sushi) = compute_gran_total_resumen(sushi_ws)
    gt_row_sushi = sushi_ws.max_row
    sushi_ws.cell(row=gt_row_sushi, column=3).value = f"{count_sushi} LOCALES"
    sushi_ws.cell(row=gt_row_sushi, column=5).value = ext_sushi
    sushi_ws.cell(row=gt_row_sushi, column=6).value = mantt_sushi
    sushi_ws.cell(row=gt_row_sushi, column=7).value = rec_sushi
    sushi_ws.cell(row=gt_row_sushi, column=8).value = cobro_sushi
    print(f"SUSHICORP GRAN TOTAL: {count_sushi} locales, {ext_sushi} ext, mantt={mantt_sushi}, rec={rec_sushi}, cobro={cobro_sushi}")

    return len(kn_data_rows), len(non_kn_data_rows)


def process_detalle(src_wb, dst_sushi_wb):
    """
    Procesa la hoja DETALLE:
    - Copia bloques KN a dst_sushi_wb
    - Elimina bloques KN de src_wb
    """
    src_ws = src_wb['DETALLE']

    # Identify KN row indices (1-based)
    kn_indices = identify_kn_rows_detalle(src_ws)

    # Create destination sheet
    sushi_ws = dst_sushi_wb.create_sheet('DETALLE')
    copy_column_dimensions(src_ws, sushi_ws)

    # Copy header row to sushi sheet
    for col_idx in range(1, src_ws.max_column + 1):
        copy_cell(src_ws.cell(row=1, column=col_idx),
                  sushi_ws.cell(row=1, column=col_idx))

    # Copy KN rows to sushi sheet (maintaining order)
    sushi_row = 2
    for row_idx in sorted(kn_indices):
        for col_idx in range(1, src_ws.max_column + 1):
            copy_cell(src_ws.cell(row=row_idx, column=col_idx),
                      sushi_ws.cell(row=sushi_row, column=col_idx))
        sushi_row += 1

    print(f"DETALLE: {len(kn_indices)} filas KN copiadas a SUSHICORP")

    # Delete KN rows from src_ws (bottom to top)
    for row_idx in sorted(kn_indices, reverse=True):
        src_ws.delete_rows(row_idx)

    print(f"DETALLE KFC: {src_ws.max_row - 1} filas de datos (sin header)")

    return len(kn_indices)


def main():
    print("=== Separando SUSHICORP de BASE_DATOS_KFC.xlsx ===\n")

    # Load source workbook
    src_wb = load_workbook(SRC)

    # Verify sheets exist
    assert 'RESUMEN_LOCALES' in src_wb.sheetnames, "Hoja RESUMEN_LOCALES no encontrada"
    assert 'DETALLE' in src_wb.sheetnames, "Hoja DETALLE no encontrada"

    # Create destination workbook for SUSHICORP
    dst_wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in dst_wb.sheetnames:
        del dst_wb['Sheet']

    # Process RESUMEN_LOCALES
    print("--- Procesando RESUMEN_LOCALES ---")
    kn_res, non_kn_res = process_resumen_locales(src_wb, dst_wb)

    # Process DETALLE
    print("\n--- Procesando DETALLE ---")
    kn_det = process_detalle(src_wb, dst_wb)

    # Save files
    print(f"\nGuardando {DST_SUSHI}...")
    dst_wb.save(DST_SUSHI)

    print(f"Guardando {SRC}...")
    src_wb.save(SRC)

    print("\n=== COMPLETADO ===")
    print(f"BASE_DATOS_KFC.xlsx    -> {non_kn_res} locales en RESUMEN_LOCALES")
    print(f"BASE_DATOS_SUSHICORP.xlsx -> {kn_res} locales en RESUMEN_LOCALES")


if __name__ == '__main__':
    main()
