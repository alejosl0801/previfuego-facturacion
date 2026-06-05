#!/usr/bin/env python3
"""
Compara cotizaciones Dropbox vs BASE_DATOS_KFC.xlsx.
Para cada INCONSISTENCIA genera detalle de líneas (cotización vs DB).
"""

import re
import io
import json
import requests
import openpyxl
from zipfile import BadZipFile
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOKEN = "sl.u.AGgAHublwFcf04LW5q-EZhcXJlLXIyxYrJUTz5TGLdhCpE5doTx3KY7AB_AbAqMh11vWqMYazpMnlwWmPzCDq_lpYGr8Iu-4mZ-_mvcDY_dT5fT4rK8wQAuYF0bBeUFMriGbLR6cXYRFVil742Aq3yn3HS-9qwM_IRS6RAEjjkJilTDvTI-wlz9103zNWf--YWbQuR9GZscifMOOO_E3puPiwLSHZ5edrlwdEOH7mgYUBFJwP4P12orInIkUbb7YRGhrlW1tH-AQ1e4_gpzES7oi6lxQqtXCsAPNYqrsxCbXj63tIPMiP-Gg7BbZaDqjDd5lypUgNV4L_A4xTgVLguiLzJrW1n93Z7ht4ccmQAw2arjH6a2AlQFrN-UjxrHwa6sIEnGp5uMUs3_SOD6Jur_ksDEsV-u933eb_2GekopgtRqJfL5f9rtluuL5PEQEKT007irY4YqwS3cdVK-t1oezX7O8AaZfPi2BmRDyk9JSJxVR0IJ7acuEQY4jP7DJCKKcawNec1EC239fPN1WBTFE9-Pwc_xljrwsoA1oBzr86NUeUEsEsltW3TnHrSavPgNaFonskH1E6AuEQs9I4UvERzEjUq83eLiCaICpWUGL8Ue13V6fxQ6drZ8praKgEIDlwZpSYdDcRsff7xcpYgqiG0rCGzpIQi2hDpSP-RaVliswad5cB_zU8vykudovR6hHXUpiLYA9L4c3ViPmyYMY8nX9vCexY9lU7wuf5yuXRwTIuk8oNgVIP6obd9jF4U92czvrcxzWDqez6n1OFWJdEd90nBv842oPPp3zx81ixwts19iZt5FtdcrOY-1VUZV3SbYPq9keEbUBlpINc2hScS-S3YF_EnE4j8rDEhmzlEoGhOqsjlRddsCs7hEytKTBat-jsxs_u5cfPHBTT03rgqeU84EFG9us9l-8pWVXcoHNLPvIuQ4KJtUi9rQ4zOluXVypolsJbMxuU5LJxc7i3H3f611E6-AulPFX0c3RCZQk11mLu5ttt5lAFp_flGId02g_HjWJUK8uZONIzvDCUm10hqlowfAwDv7ccEVmqyAf014JDpZfDt38GeY-SgTNzp7O2m0R2rAUt5-eNCiVj5eMHfuahOQfvA4DKGHmaI_Q4OzqD5kN_KbAu3RORNtOsCQfxvT485r4IU0exbXEMOoQSKfjg84mXsuQrHB-cq_ZgHL3ecmTxn6CFMjrfIaQbKtrA-Bc520qGWHFdTQgDTYDngvHtOFp2d_ySDqkOrmlV-AosU8LdP5Zyl17vH29D8BAJkd-yB3vqUUSgL2m"

DB_PATH     = "/home/user/previfuego-facturacion/BASE_DATOS_KFC.xlsx"
OUTPUT_PATH = "/home/user/previfuego-facturacion/REVISION_COTIZACIONES_2026.xlsx"
CACHE_DIR   = "/home/user/previfuego-facturacion/.cotizaciones_cache"

import os, hashlib
os.makedirs(CACHE_DIR, exist_ok=True)

MONTH_FOLDERS = [
    ("01 ENERO",      "ENERO"),
    ("02 FEBRERO",    "FEBRERO"),
    ("03 MARZO",      "MARZO"),
    ("04 ABRIL",      "ABRIL"),
    ("05 MAYO",       "MAYO"),
    ("06 JUNIO",      "JUNIO"),
    ("07 JULIO",      "JULIO"),
    ("08 AGOSTO",     "AGOSTO"),
    ("09 SEPTIEMBRE", "SEPTIEMBRE"),
    ("10 OCTUBRE",    "OCTUBRE"),
]

RECARGA_MONTHS = {"OCTUBRE", "NOVIEMBRE", "DICIEMBRE"}

# Cotizaciones que NO son mantenimiento/recarga sino VENTAS de extintores u otros
# servicios → se omiten del reporte por completo. (mes, código canónico)
VENTAS = {
    ("ENERO", "K79"),   # venta de extintores; mantenimiento real en JUNIO
    ("ABRIL", "K88"),   # venta; recarga real en OCTUBRE
}

# Archivos específicos a ignorar completamente (identificados por número de cotización).
# Son servicios puntuales fuera del ciclo mantenimiento/recarga anual.
SKIP_FILES = {
    "COTIZACION1464",   # R10 servicio independiente; la cotización del ciclo es COTIZACION1463
}


# ─── CODE NORMALISATION ──────────────────────────────────────────────────────

def canonical_code(raw: str) -> str:
    s = raw.strip().upper()
    s = re.sub(r'^KFCK', 'K', s)
    s = re.sub(r'^JV',   'V', s)
    s = re.sub(r'^CJNC', 'J', s)
    s = re.sub(r'^CN',   'BS', s)
    m = re.match(r'^([A-Z]+?)0*([1-9]\d*)', s)
    return (m.group(1) + m.group(2)) if m else s


def extract_code_from_filename(filename: str) -> str:
    name = re.sub(r'\.xlsx$', '', filename, flags=re.IGNORECASE)
    m = re.match(r'^COTIZACION\d+\s+(\S+)', name, re.IGNORECASE)
    return canonical_code(m.group(1)) if m else ""


# ─── DROPBOX ─────────────────────────────────────────────────────────────────

def dbx_list(path: str) -> list:
    cache_key  = "list_" + hashlib.md5(path.encode()).hexdigest()
    cache_file = os.path.join(CACHE_DIR, cache_key + ".json")
    if os.path.exists(cache_file):
        with open(cache_file) as f:
            return json.load(f)

    url = "https://api.dropboxapi.com/2/files/list_folder"
    hdrs = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}
    r = requests.post(url, headers=hdrs, data=json.dumps({"path": path, "limit": 2000}))
    if r.status_code != 200:
        print(f"  [WARN] list_folder failed {path}: {r.status_code}")
        return []
    result = r.json()
    entries = result.get("entries", [])
    while result.get("has_more"):
        r2 = requests.post(
            "https://api.dropboxapi.com/2/files/list_folder/continue",
            headers=hdrs,
            data=json.dumps({"cursor": result["cursor"]}),
        )
        result = r2.json()
        entries.extend(result.get("entries", []))

    with open(cache_file, "w") as f:
        json.dump(entries, f)
    return entries


def dbx_download(path: str) -> bytes:
    cache_key  = hashlib.md5(path.encode()).hexdigest()
    cache_file = os.path.join(CACHE_DIR, cache_key + ".xlsx")
    if os.path.exists(cache_file):
        with open(cache_file, "rb") as f:
            return f.read()
    r = requests.post(
        "https://content.dropboxapi.com/2/files/download",
        headers={"Authorization": f"Bearer {TOKEN}",
                 "Dropbox-API-Arg": json.dumps({"path": path})},
    )
    r.raise_for_status()
    with open(cache_file, "wb") as f:
        f.write(r.content)
    return r.content


# ─── COTIZACION PARSING ───────────────────────────────────────────────────────

_MOVIL_KEYWORDS = ("MOVIL", "VIATICO", "TRASLADO", "FLETE", "KILOMETRO", "RECORRIDO")

def parse_cotizacion(wb: openpyxl.Workbook) -> dict:
    """
    Returns {subtotal, items: [{qty,desc,unit_price,line_total}], movil_items: [{desc,total}]}
    - items: rows with col4 containing EXTINTOR
    - movil_items: rows with col4 containing MOVIL/VIATICO/etc
    """
    ws = wb.active
    items = []
    movil_items = []
    subtotal = None
    total = None

    for row in ws.iter_rows():
        cells = {c.column: c.value for c in row if c.value is not None}
        qty  = cells.get(3)
        desc = cells.get(4, "")
        unit = cells.get(10)
        ltot = cells.get(13)

        if isinstance(desc, str):
            du = desc.upper()
            if ("EXTINTOR" in du and isinstance(qty, (int, float)) and qty > 0
                    and isinstance(unit, (int, float))):
                items.append({
                    "qty": int(qty),
                    "desc": desc.strip(),
                    "unit_price": float(unit),
                    "line_total": float(ltot) if isinstance(ltot, (int, float)) else float(qty * unit),
                })
            elif any(k in du for k in _MOVIL_KEYWORDS):
                t = float(ltot) if isinstance(ltot, (int, float)) and ltot > 0 else 0.0
                if t == 0 and isinstance(qty, (int, float)) and isinstance(unit, (int, float)):
                    t = float(qty) * float(unit)
                if t > 0:
                    movil_items.append({"desc": desc.strip(), "total": t})

        # SUBTOTAL / TOTAL labels
        for c in row:
            if c.value and str(c.value).strip().upper() == "SUBTOTAL":
                for c2 in ws[c.row]:
                    if c2.column > c.column and isinstance(c2.value, (int, float)):
                        subtotal = float(c2.value)
                        break
            if c.value and str(c.value).strip().upper() in ("TOTAL", "TOTAL:"):
                for c2 in ws[c.row]:
                    if c2.column > c.column and isinstance(c2.value, (int, float)):
                        total = float(c2.value)
                        break

    if subtotal is None and total is not None:
        subtotal = round(total / 1.15, 2)

    return {"subtotal": subtotal, "items": items, "movil_items": movil_items}


def fmt_items(items: list, movil_items: list = None, price_key: str = "unit_price") -> str:
    """Format line items as '2×50lbCO2@$20 + 1×10lbPQS@$4 [+movil:$40]'."""
    parts = []
    for it in items:
        desc_short = shorten_desc(it["desc"])
        parts.append(f"{it['qty']}×{desc_short}@${it[price_key]:.0f}")
    base = " + ".join(parts) if parts else "(sin líneas ext)"
    if movil_items:
        movil_total = sum(m["total"] for m in movil_items)
        movil_desc  = "; ".join(m["desc"] for m in movil_items)
        base += f"  [+MOVIL:${movil_total:.2f} — {movil_desc}]"
    return base


def shorten_desc(desc: str) -> str:
    """Convert 'EXTINTOR 50 LIBRAS - GAS CARBONICO' → '50lbCO2'."""
    d = desc.upper()
    # Extract capacity
    cap_m = re.search(r'(\d[\d,\.]*)\s*(LIBRAS?|GLNS?|KGS?|LBS?)', d)
    cap = cap_m.group(0).replace(" ", "").replace("LIBRAS", "lb").replace("LIBRA", "lb") \
                        .replace("GLNS", "gln").replace("GLN", "gln") \
                        .replace("KGS", "kg").replace("LBS", "lb") if cap_m else "?"
    cap = re.sub(r'[,]', '.', cap)
    # Extract type
    if "CARBONICO" in d or "CO2" in d or "DIÓXIDO" in d:
        tipo = "CO2"
    elif "POLVO" in d or "PQS" in d or "ABC" in d:
        tipo = "PQS"
    elif "POTASIO" in d or "ACETATO" in d or "CLASE K" in d or "COCINA" in d:
        tipo = "K"
    elif "HALOTRON" in d or "HALON" in d:
        tipo = "HAL"
    elif "AGUA" in d:
        tipo = "H2O"
    else:
        tipo = "?"
    return f"{cap}{tipo}"


# ─── DB LOADING ───────────────────────────────────────────────────────────────

def load_db_full(db_path: str):
    """
    Returns:
      resumen:  dict  code → {codigo, nombre, mes, mantt, recarga, cobro, n_ext, ano_recarga}
      detalle:  dict  code → list of extintor rows {tipo, cap, ubic, mantt, recarga}
      movil_db: dict  code → {desc, mantt, recarga}  (MOVILIZACIÓN rows only)
    """
    wb = openpyxl.load_workbook(db_path)

    # RESUMEN_LOCALES
    ws_res = wb["RESUMEN_LOCALES"]
    resumen = {}
    for row in ws_res.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        codigo = str(row[0])
        canon = canonical_code(codigo)
        resumen[canon] = {
            "codigo": codigo,
            "nombre": str(row[2]) if row[2] else "",
            "mes": str(row[3]).upper() if row[3] else "",
            "n_ext": row[4],
            "mantt":   float(row[5]) if row[5] is not None else None,
            "recarga": float(row[6]) if row[6] is not None else None,
            "cobro":   float(row[7]) if row[7] is not None else None,
            "ano_recarga": int(row[9]) if row[9] is not None else None,
        }

    # DETALLE — split extintor rows from MOVILIZACIÓN rows
    ws_det   = wb["DETALLE"]
    detalle  = defaultdict(list)
    movil_db = {}
    for row in ws_det.iter_rows(min_row=2, values_only=True):
        nombre_local = row[1]
        if not nombre_local or not isinstance(nombre_local, str):
            continue
        if "TOTALES" in nombre_local.upper():
            continue
        m = re.match(r'^([A-Z0-9]+)\s*[-–]', nombre_local.strip())
        if not m:
            continue
        canon = canonical_code(m.group(1))
        tipo_row = str(row[4]).upper() if row[4] else ""
        if tipo_row == "MOVILIZACIÓN":
            movil_db[canon] = {
                "desc":    str(row[3]) if row[3] else "",
                "mantt":   float(row[6]) if row[6] is not None else 0,
                "recarga": float(row[7]) if row[7] is not None else 0,
            }
        else:
            detalle[canon].append({
                "tipo":    str(row[4]) if row[4] else "",
                "cap":     str(row[5]) if row[5] else "",
                "ubic":    str(row[3]) if row[3] else "",
                "mantt":   float(row[6]) if row[6] is not None else 0,
                "recarga": float(row[7]) if row[7] is not None else 0,
            })

    return resumen, detalle, movil_db


def _cap_short(cap: str, tipo: str) -> str:
    """Normalize cap+tipo for comparison key, e.g. '10 LBS' + 'PQS' → '10lbPQS'."""
    c = re.sub(r'\s+', '',
               cap.replace(" LBS","lb").replace(" LB","lb")
                  .replace(" GLNS","gln").replace(" GLN","gln")
                  .replace(" KGS","kg").replace(" KG","kg"))
    t = {"CO2":"CO2","PQS":"PQS","K":"K","HALOTRON":"HAL"}.get(tipo.upper(), tipo)
    return f"{c}{t}"


def fmt_db_items(db_exts: list, tipo_cobro: str, movil: dict = None) -> str:
    """Format DB extintores (excl. MOVILIZACIÓN) as '1×10lbPQS@$4 ...' optionally appending movil."""
    if not db_exts and not movil:
        return "(vacío en DB)"
    parts = []
    for e in db_exts:
        cs = _cap_short(e["cap"], e["tipo"])
        price = e["recarga"] if tipo_cobro == "RECARGA" else e["mantt"]
        parts.append(f"1×{cs}@${price:.0f}")
    base = " + ".join(parts) if parts else "(sin ext en DB)"
    if movil and movil.get("mantt", 0) > 0:
        mv = movil["recarga"] if tipo_cobro == "RECARGA" else movil["mantt"]
        base += f"  [+MOVIL:${mv:.2f} — {movil['desc']}]"
    return base


def build_analysis(cot_items: list, cot_movil: list, db_exts: list, db_movil: dict,
                   subtotal_cot: float, valor_db: float, tipo_cobro: str) -> tuple:
    """
    Returns (analisis: str, recomendacion: str)
    Compares cotización ext lines vs DB ext rows (viáticos separated).
    """
    price_key   = "recarga" if tipo_cobro == "RECARGA" else "mantt"
    movil_cot   = sum(m["total"] for m in cot_movil) if cot_movil else 0.0
    movil_db_v  = (db_movil[price_key] if db_movil else 0.0) if db_movil else 0.0

    sum_ext_cot = sum(it["qty"] * it["unit_price"] for it in cot_items)
    sum_ext_db  = sum(e[price_key] for e in db_exts)

    # If movil not detected in cot lines but subtotal > sum_ext, infer from difference
    if movil_cot == 0 and movil_db_v > 0 and subtotal_cot:
        inferred = round(subtotal_cot - sum_ext_cot, 2)
        if abs(inferred - movil_db_v) < 1.0:
            movil_cot = inferred

    n_cot = sum(it["qty"] for it in cot_items)
    n_db  = len(db_exts)

    # Per-type counts
    cot_by_type: dict = defaultdict(lambda: {"qty": 0, "unit": 0.0})
    for it in cot_items:
        t = shorten_desc(it["desc"])
        cot_by_type[t]["qty"]  += it["qty"]
        cot_by_type[t]["unit"]  = it["unit_price"]

    db_by_type: dict = defaultdict(lambda: {"qty": 0, "unit": 0.0})
    for e in db_exts:
        t = _cap_short(e["cap"], e["tipo"])
        db_by_type[t]["qty"]  += 1
        db_by_type[t]["unit"]  = e[price_key]

    # Build extras / missing lists
    all_types = set(cot_by_type) | set(db_by_type)
    extras, missing, changed = [], [], []
    for t in all_types:
        cq = cot_by_type[t]["qty"]
        dq = db_by_type[t]["qty"]
        cu = cot_by_type[t]["unit"]
        du = db_by_type[t]["unit"]
        if cq > dq:
            extras.append((t, cq - dq, cu))
        elif dq > cq:
            missing.append((t, dq - cq, du))
        elif abs(cu - du) > 0.5 and cq > 0:
            changed.append((t, cq, cu, du))

    parts_a = []  # ANALISIS parts
    parts_r = []  # RECOMENDACIÓN parts

    # ── Viáticos comparison ────────────────────────────────────────────────
    if movil_db_v > 0 and abs(movil_cot - movil_db_v) > 0.5:
        parts_a.append(f"Viáticos COT=${movil_cot:.2f} vs DB=${movil_db_v:.2f}")
        if movil_cot > 0:
            parts_r.append(f"Actualizar viáticos en DB: ${movil_db_v:.2f}→${movil_cot:.2f}")
        else:
            parts_a.append("(viáticos no detectados en líneas COT, posible formato)")

    # ── Extintor net comparison ────────────────────────────────────────────
    net_diff = round(sum_ext_cot - sum_ext_db, 2)
    if abs(net_diff) > 0.5:
        sign = "+" if net_diff > 0 else ""
        parts_a.append(
            f"Extintores COT=${sum_ext_cot:.2f} vs DB=${sum_ext_db:.2f} "
            f"(neto {sign}${net_diff:.2f})"
        )

    # ── Count diff ────────────────────────────────────────────────────────
    if n_cot != n_db:
        parts_a.append(f"Cantidad: {n_cot} ext en COT vs {n_db} en DB")

    # ── Per-type details ──────────────────────────────────────────────────
    for t, qty, unit in extras:
        parts_a.append(f"COT tiene +{qty}×{t} que no está en DB (precio ${unit:.0f}/ext)")
        parts_r.append(f"¿AGREGAR {qty}×{t} a BD? (${unit:.0f}/ext)")
    for t, qty, unit in missing:
        parts_a.append(f"DB tiene +{qty}×{t} que no aparece en COT (${unit:.0f}/ext)")
        parts_r.append(f"¿RETIRAR {qty}×{t} de BD? (o verificar si fue omitido en COT)")
    for t, qty, cu, du in changed:
        parts_a.append(f"{qty}×{t}: precio COT=${cu:.0f} vs DB=${du:.0f}/ext")
        parts_r.append(f"¿Actualizar precio {t} en BD: ${du:.0f}→${cu:.0f}/ext?")

    # ── Large unexplained gap ─────────────────────────────────────────────
    if not extras and not missing and not changed:
        if abs(net_diff) > 0.5:
            parts_a.append("Mismo número y tipo de extintores pero precio diferente")
        elif abs(sum_ext_cot - (subtotal_cot or 0)) > 0.5 and movil_cot == 0:
            unexp = round((subtotal_cot or 0) - sum_ext_cot, 2)
            parts_a.append(
                f"⚠ Subtotal ${subtotal_cot:.2f} incluye ${unexp:.2f} no identificado "
                f"(viáticos u otro cargo sin línea reconocida)"
            )
            parts_r.append("Revisar cotización: hay cargo no identificado en subtotal")

    analisis = " | ".join(parts_a) if parts_a else "Sin diferencias por tipo (revisar precio unitario)"
    recom    = " | ".join(parts_r) if parts_r else "—"
    return analisis, recom


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("Cargando base de datos...")
    resumen, detalle, movil_db = load_db_full(DB_PATH)
    print(f"  {len(resumen)} locales en RESUMEN, {sum(len(v) for v in detalle.values())} filas en DETALLE, "
          f"{len(movil_db)} con movilización")

    rows = []

    for folder_name, mes_name in MONTH_FOLDERS:
        dropbox_path = f"/Previfuego/2026/{folder_name}/GRUPO KFC"
        print(f"\n{'='*60}\nProcesando: {folder_name}")

        entries = dbx_list(dropbox_path)
        if not entries:
            print(f"  [WARN] vacío o no encontrado: {dropbox_path}")
            continue

        xlsx_files = [
            e for e in entries
            if e[".tag"] == "file"
            and e["name"].lower().endswith(".xlsx")
            and e["name"].upper().startswith("COTIZACION")
        ]
        print(f"  {len(xlsx_files)} COTIZACION*.xlsx")

        for entry in xlsx_files:
            filename  = entry["name"]
            file_path = entry["path_lower"]
            local_code = extract_code_from_filename(filename)

            # Skip specific cotizaciones (servicios fuera del ciclo mantt/recarga)
            cot_num_m = re.match(r'^(COTIZACION\d+)', filename, re.IGNORECASE)
            if cot_num_m and cot_num_m.group(1).upper() in SKIP_FILES:
                print(f"  [{filename}] → IGNORADO (fuera de ciclo)")
                continue

            print(f"  [{filename}] → {local_code!r}", end="")

            # Download
            try:
                content = dbx_download(file_path)
            except Exception as e:
                print(f" → DOWNLOAD ERROR: {e}")
                rows.append(_row(mes_name, filename, local_code, "", None, None, "", None, "ERROR",
                                 "", "", f"Download: {e}"))
                continue

            # Parse xlsx
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            except BadZipFile:
                print(" → CORRUPTED")
                rows.append(_row(mes_name, filename, local_code, "", None, None, "", None, "ERROR",
                                 "", "", "Archivo corrupto"))
                continue
            except Exception as e:
                print(f" → PARSE ERROR: {e}")
                rows.append(_row(mes_name, filename, local_code, "", None, None, "", None, "ERROR",
                                 "", "", str(e)))
                continue

            parsed     = parse_cotizacion(wb)
            subtotal   = parsed["subtotal"]
            cot_items  = parsed["items"]
            cot_movil  = parsed["movil_items"]

            # Cotización de VENTA → omitir completamente del reporte
            if (mes_name, local_code) in VENTAS:
                print(f" → VENTA/omitido (subtotal={subtotal})")
                continue

            # DB lookup
            db_entry = resumen.get(local_code)
            if db_entry is None:
                print(f" → SIN MATCH DB (subtotal={subtotal})")
                rows.append(_row(mes_name, filename, local_code, "", subtotal, None, "",
                                 None, "SIN_MATCH_DB",
                                 fmt_items(cot_items, cot_movil), "", "", ""))
                continue

            nombre   = db_entry["nombre"]
            db_exts  = detalle.get(local_code, [])
            db_movil = movil_db.get(local_code)

            # Determine tipo (RECARGA vs MANTENIMIENTO)
            if db_entry.get("ano_recarga") == 2026 and mes_name not in RECARGA_MONTHS:
                tipo = "RECARGA"
            else:
                tipo = "RECARGA" if mes_name in RECARGA_MONTHS else "MANTENIMIENTO"
            # For RECARGA months use COBRO_ANUAL (handles mixed-service locals like B1
            # where some extintores get MANTT while others get RECARGA in the same month).
            # For MANTT months, COBRO_ANUAL = TOTAL_RECARGA ≠ TOTAL_MANTT, so keep mantt.
            valor_db_v = db_entry["cobro"] if tipo == "RECARGA" else db_entry["mantt"]

            # Status
            if subtotal is None:
                estado = "SIN_SUBTOTAL"; diff_pct = None
            elif not valor_db_v:
                estado = "SIN_VALOR_DB"; diff_pct = None
            else:
                diff_pct = abs(subtotal - valor_db_v) / valor_db_v * 100
                if diff_pct <= 1.0:
                    estado = "OK"
                elif tipo == "RECARGA" and db_entry["mantt"]:
                    diff_vs_mantt = abs(subtotal - db_entry["mantt"]) / db_entry["mantt"] * 100
                    estado = "COT_DEBE_RECARGA" if diff_vs_mantt <= 1.0 else "INCONSISTENCIA"
                else:
                    estado = "INCONSISTENCIA"

            print(f" → {estado}" + (f" diff={diff_pct:.1f}%" if diff_pct else ""))

            cot_det = db_det = analisis = recom = ""
            if estado in ("INCONSISTENCIA", "COT_DEBE_RECARGA"):
                cot_det = fmt_items(cot_items, cot_movil)
                db_det  = fmt_db_items(db_exts, tipo, db_movil)
                analisis, recom = build_analysis(
                    cot_items, cot_movil, db_exts, db_movil,
                    subtotal, valor_db_v or 0, tipo
                )
                if estado == "COT_DEBE_RECARGA":
                    analisis = (
                        f"⚠ COT usa precio MANTENIMIENTO (${subtotal:.2f}). "
                        f"Debería ser RECARGA (${valor_db_v:.2f}). | {analisis}"
                    ).rstrip(" |")
                    recom = f"Reemitir cotización como RECARGA por ${valor_db_v:.2f}"

            rows.append(_row(mes_name, filename, local_code, nombre,
                             subtotal, valor_db_v, tipo,
                             round(diff_pct, 2) if diff_pct is not None else None,
                             estado, cot_det, db_det, analisis, recom))

    # ─── EXCEL ────────────────────────────────────────────────────────────────
    print(f"\n{'='*60}\nGenerando Excel…")
    _write_excel(rows)

    # ─── PRINT SUMMARY ────────────────────────────────────────────────────────
    total = len(rows)
    ok              = sum(1 for r in rows if r["ESTADO"] == "OK")
    inconsistencias = sum(1 for r in rows if r["ESTADO"] == "INCONSISTENCIA")
    cot_rec         = sum(1 for r in rows if r["ESTADO"] == "COT_DEBE_RECARGA")
    sin_match       = sum(1 for r in rows if r["ESTADO"] == "SIN_MATCH_DB")
    errores         = sum(1 for r in rows if r["ESTADO"] == "ERROR")

    print(f"\n{'='*60}")
    print(f"  Total revisados  : {total}")
    print(f"  OK               : {ok}")
    print(f"  INCONSISTENCIAS  : {inconsistencias}")
    print(f"  COT_DEBE_RECARGA : {cot_rec}")
    print(f"  SIN MATCH DB     : {sin_match}")
    print(f"  ERRORES          : {errores}")

    print("\nINCONSISTENCIAS:")
    for r in rows:
        if r["ESTADO"] == "INCONSISTENCIA":
            print(f"  {r['MES']:12} {r['LOCAL_CODE']:8} {str(r['LOCAL_NOMBRE'])[:32]:32} "
                  f"COT=${r['SUBTOTAL_COT']:7.2f} DB=${r['VALOR_DB']:7.2f} "
                  f"({r['DIFERENCIA%']:.1f}%) | {r['ANALISIS'][:80]}")


# ─── HELPERS ─────────────────────────────────────────────────────────────────

HEADERS = ["MES", "ARCHIVO", "LOCAL_CODE", "LOCAL_NOMBRE",
           "SUBTOTAL_COT", "VALOR_DB", "TIPO", "DIFERENCIA%",
           "ESTADO", "COT_DETALLE", "DB_DETALLE", "ANALISIS", "RECOMENDACIÓN"]


def _row(mes, arch, code, nombre, subtot, val_db, tipo, diff, estado,
         cot_det, db_det, analisis, recom=""):
    return {
        "MES": mes, "ARCHIVO": arch, "LOCAL_CODE": code, "LOCAL_NOMBRE": nombre,
        "SUBTOTAL_COT": subtot, "VALOR_DB": val_db, "TIPO": tipo,
        "DIFERENCIA%": diff, "ESTADO": estado,
        "COT_DETALLE": cot_det, "DB_DETALLE": db_det,
        "ANALISIS": analisis, "RECOMENDACIÓN": recom,
    }


def _write_excel(rows: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "REVISION"

    thin   = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")

    ws.append(HEADERS)
    for ci in range(1, len(HEADERS) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    fill_ok    = PatternFill("solid", fgColor="C6EFCE")
    fill_inc   = PatternFill("solid", fgColor="FFC7CE")
    fill_rec   = PatternFill("solid", fgColor="F4B942")
    fill_warn  = PatternFill("solid", fgColor="FFEB9C")
    fill_venta = PatternFill("solid", fgColor="BDD7EE")
    fill_err   = PatternFill("solid", fgColor="D9D9D9")

    for r in rows:
        ws.append([r[h] for h in HEADERS])
        rn = ws.max_row
        est = r["ESTADO"]
        fill = (fill_ok  if est == "OK"            else
                fill_inc if est == "INCONSISTENCIA" else
                fill_rec if est == "COT_DEBE_RECARGA" else
                fill_venta if est == "VENTA" else
                fill_warn if est in ("SIN_MATCH_DB","SIN_SUBTOTAL","SIN_VALOR_DB") else
                fill_err)
        for ci in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=rn, column=ci)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(ci >= 10))

    # Column widths
    for ci, w in enumerate([10, 52, 12, 38, 12, 10, 14, 11, 18, 62, 62, 85, 70], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── RESUMEN sheet ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("RESUMEN")
    total = len(rows)
    ok   = sum(1 for r in rows if r["ESTADO"] == "OK")
    inc  = sum(1 for r in rows if r["ESTADO"] == "INCONSISTENCIA")
    rec  = sum(1 for r in rows if r["ESTADO"] == "COT_DEBE_RECARGA")
    sm   = sum(1 for r in rows if r["ESTADO"] == "SIN_MATCH_DB")
    err  = sum(1 for r in rows if r["ESTADO"] == "ERROR")

    for sd in [
        ["RESUMEN REVISION COTIZACIONES 2026", ""],
        ["", ""],
        ["Total archivos revisados",                    total],
        ["✅ OK (≤1% diferencia)",                       ok],
        ["🔴 INCONSISTENCIA de precio (>1%)",            inc],
        ["🟠 COT_DEBE_RECARGA (emitida como MANT/OCT)", rec],
        ["🟡 SIN MATCH en DB (otras marcas)",            sm],
        ["⚫ ERRORES (descarga/parseo)",                  err],
    ]:
        ws2.append(sd)
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 48
    ws2.column_dimensions["B"].width = 12

    ws2.append([""])
    ws2.append(["DETALLE POR MES", "TOTAL", "OK", "INCONSISTENCIA", "COT_DEBE_RECARGA", "SIN_MATCH", "OTROS"])
    by_month = defaultdict(list)
    for r in rows:
        by_month[r["MES"]].append(r)
    for _, mes in MONTH_FOLDERS:
        mr = by_month[mes]
        if not mr:
            continue
        m_ok  = sum(1 for r in mr if r["ESTADO"] == "OK")
        m_inc = sum(1 for r in mr if r["ESTADO"] == "INCONSISTENCIA")
        m_rec = sum(1 for r in mr if r["ESTADO"] == "COT_DEBE_RECARGA")
        m_sm  = sum(1 for r in mr if r["ESTADO"] == "SIN_MATCH_DB")
        m_oth = len(mr) - m_ok - m_inc - m_rec - m_sm
        ws2.append([mes, len(mr), m_ok, m_inc, m_rec, m_sm, m_oth])

    wb.save(OUTPUT_PATH)
    print(f"Guardado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
